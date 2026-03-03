"""Compute marginal costs from Orange & Rockland's MCOS study workbook.

Same NERA methodology as ConEd (sister companies), but with different cost
center structure and workbook layout.  Produces four MC variants × two formats
= 8 output CSVs.

O&R's CapEx Transmission sheet contains three 138 kV projects.  We split them
into bulk TX (excluded) and local TX (included in sub-TX + dist):

  Bulk TX (excluded):
    - West Nyack: new 138 kV UG line, Burns→West Nyack.  Confirmed in NYISO
      Gold Book Table VII (line 3962, p. 161).

  Local TX (included as sub-TX):
    - Oak Street: reconductor existing line to standard 138 kV.  NOT in Gold
      Book — would be dropped under a "Gold Book = bulk TX" approach.
    - New Hempstead: reconductor existing line to standard 138 kV.  NOT in
      Gold Book — same reasoning.

  Justification for reclassifying Oak St. and New Hempstead as sub-TX:
    1. Not in NYISO Gold Book Table VII → not NYISO-jurisdictional bulk TX
    2. 138 kV reconductoring of existing local lines, not new backbone capacity
    3. Without reclassification, $36.5M in capital falls through the gap
       (excluded from our analysis AND absent from the bulk TX analysis)
    4. Small scale ($29M + $7.5M) serving O&R's Eastern NY load area

The remaining cost centers map cleanly:
  - Area Station & Sub-TX (CapEx Substation) — cumulative, INCLUDED
  - Primary (CapEx Primary) — cumulative, INCLUDED
  - Secondary Distribution (CapEx Secondary) — flat $/kW, INCLUDED

MC formula: same as ConEd — see analyze_coned_mcos.py docstring.
Local TX uses the Transmission composite rate (same plant type / carrying
charge characteristics).

Outputs (8 CSVs):
  - or_{cumulative,incremental}_{diluted,undiluted}_{levelized,annualized}.csv
  - Terminal report

Usage (via Justfile):
    just analyze-or
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import fsspec
import openpyxl
import polars as pl


# ── Study parameters ─────────────────────────────────────────────────────────

YEARS = list(range(2025, 2035))
N_YEARS = len(YEARS)
LEVELIZATION_YEARS = range(2026, 2033)  # 7-year window for BAT input
REBASE_YEAR = 2026  # all real values rebased to 2026 dollars

COST_CENTERS = ["transmission", "tx_local", "substation", "primary", "secondary_dist"]
LOCAL_CENTERS = ["tx_local", "substation", "primary", "secondary_dist"]

CC_LABELS = {
    "transmission": "Bulk TX (Gold Book, 138 kV)",
    "tx_local": "Local TX (non-Gold-Book, 138 kV)",
    "substation": "Area Station & Sub-TX",
    "primary": "Primary Distribution",
    "secondary_dist": "Secondary Distribution",
    "local_total": "Sub-TX + dist (excl. bulk TX)",
}

VARIANT_NAMES = [
    "cumulative_diluted",
    "incremental_diluted",
    "cumulative_undiluted",
    "incremental_undiluted",
]


# ── O&R workbook cell references ─────────────────────────────────────────────
#
# Verified from direct inspection of or_study_workpaper.xlsx.

SH_TX = "CapEx Transmission"
SH_SUB = "CapEx Substation"
SH_PRI = "CapEx Primary"
SH_SEC = "CapEx Secondary"
SH_SCHED10 = "Carrying Charge Loaders"
SH_COINC = "Coincident Forecast"

# ── CapEx Transmission / Substation: column layout ───────────────────────────
#
# Left half:
#   B(2)=$ (000s)  C(3)=MW  D(4)=$/kW  E(5)=Region  F(6)=Area Substation
#   G(7)..P(16) = years 2025–2034
#
# Right half:
#   R(18)=Location  S(19)=Description  T(20)=MW  U(21)=CWE ($000s)
#   V(22)=Total Un-Escalated  W(23)..AF(32) = cumulative cashflow 2025–2034

CAPEX_YEAR_COL = {yr: 7 + (yr - 2025) for yr in YEARS}  # G(7)=2025 .. P(16)=2034
CAPEX_MW_COL = 3  # C: MW (left half)
CAPEX_DESC_COL = 19  # S: Description (right half)
CAPEX_COST_COL = 21  # U: CWE $000s (right half)
CAPEX_RIGHT_MW = 20  # T: MW (right half)

# CapEx Transmission project rows — split into bulk and local.
TX_BULK_ROWS = [8]  # West Nyack (Gold Book line 3962)
TX_LOCAL_ROWS = [9, 10]  # Oak St., New Hempstead (not in Gold Book)
TX_ALL_ROWS = range(8, 11)

# Right-half cumulative cashflow columns ($000s).
CAPEX_CF_YEAR_COL = {yr: 23 + (yr - 2025) for yr in YEARS}  # W(23)=2025 .. AF(32)=2034
SUB_PROJECT_ROWS = range(
    8, 12
)  # rows 8–11 (4 projects; row 12 is a total, not a project)

# ── CapEx Primary ────────────────────────────────────────────────────────────
PRI_REGION_ROWS = [57, 58, 59]  # Central, Eastern, Western ($000s)
PRI_PROJECT_ROWS = range(8, 34)  # rows 8–33 (26 projects, right half)
PRI_YEAR_COL = CAPEX_YEAR_COL  # same G(7)..P(16)

# ── CapEx Secondary ──────────────────────────────────────────────────────────
# Row 18, col F(6): "Secondary Distribution ($/kW)" = 12.5659
# This is capital cost per kW of SYSTEM PEAK, not per kW of project capacity.
SEC_RATE_ROW = 18
SEC_RATE_COL = 6  # F

# ── Schedule 10: Carrying Charge Loaders ─────────────────────────────────────
SCHED10_RATE_COL = 15  # Col O = "Annual MC (System Peak)"
SCHED10_ROWS = {
    "transmission": 12,
    "tx_local": 12,  # same carrying charge as transmission (same plant type)
    "substation": 13,
    "primary": 14,
    "secondary_dist": 17,  # "Secondary Distribution" (combined; rows 15-16 marked "remove")
}
SCHED10_ESC_ROW = 26
SCHED10_ESC_YEAR_COL = {yr: 3 + (yr - 2025) for yr in YEARS}

# ── Coincident Forecast ──────────────────────────────────────────────────────
COINC_TOTAL_ROW = 65
COINC_YEAR_COL = {yr: 5 + (yr - 2025) for yr in YEARS}
COINC_2024_COL = 4


# ── Data classes ─────────────────────────────────────────────────────────────


@dataclass
class ProjectData:
    """Per-project data parsed from right-half cashflow columns."""

    name: str
    capacity_mw: float
    final_capital_k: float  # cumulative capital at completion ($000s), or annual budget
    in_service_year: int  # first year where cashflow stabilizes, or first nonzero year


@dataclass
class CostCenterData:
    name: str
    label: str
    n_projects: int
    total_capacity_mw: float
    total_cost_k: float
    composite_rate: float
    capital_by_year: dict[int, float] = field(default_factory=dict)
    capacity_by_year: dict[int, float] = field(default_factory=dict)
    is_annual: bool = False
    is_per_kw_system: bool = False  # True for secondary_dist ($/kW of system peak)
    projects: list[ProjectData] = field(default_factory=list)


@dataclass
class MCRow:
    year: int
    capital_k: float
    annual_rr_k: float
    escalation: float
    nominal_mc: float
    real_mc: float


# ── Workbook I/O ─────────────────────────────────────────────────────────────


def _cell(sheet, row: int, col: int) -> float:
    v = sheet.cell(row=row, column=col).value
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _str_cell(sheet, row: int, col: int) -> str:
    v = sheet.cell(row=row, column=col).value
    return str(v).strip() if v else ""


def _load_wb(path: str):
    if path.startswith("s3://"):
        fs = fsspec.filesystem("s3")
        with fs.open(path, "rb") as f:
            return openpyxl.load_workbook(f, data_only=True)
    return openpyxl.load_workbook(path, data_only=True)


# ── Parsing ──────────────────────────────────────────────────────────────────


def parse_composite_rates(wb) -> dict[str, float]:
    sheet = wb[SH_SCHED10]
    return {
        name: _cell(sheet, row, SCHED10_RATE_COL) for name, row in SCHED10_ROWS.items()
    }


def parse_escalation(wb) -> dict[int, float]:
    sheet = wb[SH_SCHED10]
    return {
        yr: _cell(sheet, SCHED10_ESC_ROW, col)
        for yr, col in SCHED10_ESC_YEAR_COL.items()
    }


def parse_coincident_peak(wb) -> dict[int, float]:
    sheet = wb[SH_COINC]
    peaks = {
        yr: _cell(sheet, COINC_TOTAL_ROW, col) for yr, col in COINC_YEAR_COL.items()
    }
    peaks[2024] = _cell(sheet, COINC_TOTAL_ROW, COINC_2024_COL)
    return peaks


def _detect_in_service_year(cashflows: list[float]) -> int:
    """Infer in-service year from cumulative cashflow stabilization.

    Returns the first study year where the cashflow equals its final value,
    indicating CWIP has ended and the project is complete.
    Falls back to the last study year if cashflow never fully stabilizes.
    """
    final = cashflows[-1]
    if final <= 0:
        return YEARS[-1]
    for i, v in enumerate(cashflows):
        if abs(v - final) < 0.5:
            return YEARS[i]
    return YEARS[-1]


def _parse_projects_from_cashflow(
    sheet,
    rows: list[int] | range,
    cf_cols: dict[int, int],
    mw_col: int,
    name_col: int,
) -> list[ProjectData]:
    """Read per-project cashflow rows, detect in-service years."""
    projects: list[ProjectData] = []
    for r in rows:
        ref = _str_cell(sheet, r, name_col)
        if not ref:
            continue
        mw = _cell(sheet, r, mw_col)
        cashflows = [_cell(sheet, r, col) for col in cf_cols.values()]
        in_svc = _detect_in_service_year(cashflows)
        projects.append(ProjectData(ref, mw, cashflows[-1], in_svc))
    return projects


def _build_scoped_capital_capacity(
    projects: list[ProjectData],
) -> tuple[dict[int, float], dict[int, float]]:
    """Build capital_by_year and capacity_by_year using in-service-year scoping."""
    capital: dict[int, float] = {}
    capacity: dict[int, float] = {}
    for yr in YEARS:
        in_scope = [p for p in projects if p.in_service_year <= yr]
        capital[yr] = sum(p.final_capital_k for p in in_scope)
        capacity[yr] = sum(p.capacity_mw for p in in_scope)
    return capital, capacity


def _parse_tx_rows(wb, rows: list[int], name: str) -> CostCenterData:
    """Parse a subset of CapEx Transmission project rows.

    Uses per-project cashflow columns (W-AF) with in-service-year scoping,
    matching the NiMo/CenHud methodology.
    """
    sheet = wb[SH_TX]
    projects = _parse_projects_from_cashflow(
        sheet, rows, CAPEX_CF_YEAR_COL, CAPEX_RIGHT_MW, CAPEX_DESC_COL
    )
    capital, capacity = _build_scoped_capital_capacity(projects)

    return CostCenterData(
        name=name,
        label=CC_LABELS[name],
        n_projects=len({p.name for p in projects}),
        total_capacity_mw=sum(p.capacity_mw for p in projects),
        total_cost_k=sum(p.final_capital_k for p in projects),
        composite_rate=0.0,
        capital_by_year=capital,
        capacity_by_year=capacity,
        projects=projects,
    )


def parse_capex_tx_split(wb) -> tuple[CostCenterData, CostCenterData]:
    """Split CapEx Transmission into bulk TX (Gold Book) and local TX."""
    bulk = _parse_tx_rows(wb, TX_BULK_ROWS, "transmission")
    local = _parse_tx_rows(wb, TX_LOCAL_ROWS, "tx_local")
    return bulk, local


def parse_capex_sub(wb) -> CostCenterData:
    """Parse CapEx Substation using per-project cashflow with in-service scoping."""
    sheet = wb[SH_SUB]
    projects = _parse_projects_from_cashflow(
        sheet, SUB_PROJECT_ROWS, CAPEX_CF_YEAR_COL, CAPEX_RIGHT_MW, CAPEX_DESC_COL
    )
    capital, capacity = _build_scoped_capital_capacity(projects)

    return CostCenterData(
        name="substation",
        label=CC_LABELS["substation"],
        n_projects=len({p.name for p in projects}),
        total_capacity_mw=sum(p.capacity_mw for p in projects),
        total_cost_k=sum(p.final_capital_k for p in projects),
        composite_rate=0.0,
        capital_by_year=capital,
        capacity_by_year=capacity,
        projects=projects,
    )


def parse_capex_primary(wb) -> CostCenterData:
    """Parse CapEx Primary — project-level annual budgets with in-service years.

    CapEx Primary has a DIFFERENT right-side column layout than TX/Sub:
      R(18)=Region  S(19)=Location  T(20)=Description  U(21)=MW
      V(22)=$/kW  W(23)=Budget/CWE ($000s)  X(24)..AG(33)=yearly budget 2025-2034

    Each project has a constant annual budget that starts at its in-service
    year (the first year with a non-zero value).  capital_by_year(Y) = sum
    of annual budgets for all projects in-service by Y, and capacity_by_year(Y)
    = sum of project MW for those same projects.
    """
    sheet = wb[SH_PRI]
    PRI_MW_COL = 21  # U: MW
    PRI_BUDGET_COL = 23  # W: Budget/CWE ($000s)
    PRI_YEARLY_COL = {yr: 24 + (yr - 2025) for yr in YEARS}  # X(24)=2025..AG(33)=2034

    projects: list[ProjectData] = []
    for r in PRI_PROJECT_ROWS:
        region = _str_cell(sheet, r, 18)  # R: Region
        if not region:
            continue
        loc = _str_cell(sheet, r, 19)  # S: Location
        mw = _cell(sheet, r, PRI_MW_COL)
        budget = _cell(sheet, r, PRI_BUDGET_COL)
        yearly = [_cell(sheet, r, col) for col in PRI_YEARLY_COL.values()]

        # In-service year = first year with non-zero annual budget
        in_svc = YEARS[0]
        for i, v in enumerate(yearly):
            if v > 0:
                in_svc = YEARS[i]
                break

        name = f"{region[:3]}/{loc}" if loc else region
        projects.append(ProjectData(name, mw, budget, in_svc))

    capital: dict[int, float] = {}
    capacity: dict[int, float] = {}
    for yr in YEARS:
        in_scope = [p for p in projects if p.in_service_year <= yr]
        capital[yr] = sum(p.final_capital_k for p in in_scope)
        capacity[yr] = sum(p.capacity_mw for p in in_scope)

    return CostCenterData(
        name="primary",
        label=CC_LABELS["primary"],
        n_projects=len(projects),
        total_capacity_mw=sum(p.capacity_mw for p in projects),
        total_cost_k=sum(p.final_capital_k for p in projects),
        composite_rate=0.0,
        capital_by_year=capital,
        capacity_by_year=capacity,
        projects=projects,
    )


def parse_capex_secondary(wb, system_peak_mw: float) -> CostCenterData:
    """Parse CapEx Secondary — flat $/kW for the whole system.

    The workbook gives capital cost per kW of system peak ($12.57/kW).
    We store:
      - capital_by_year: total system capital ($000s) = $/kW × system_peak_kw / 1000
      - capacity_by_year: system peak (MW) — the denominator for this cost center
    This means diluted MC = capital × rate × esc / system_peak, and
    undiluted MC = same value (since the $/kW is already per kW of system peak,
    there's no separate project capacity for undiluted).
    """
    sheet = wb[SH_SEC]
    dollars_per_kw = _cell(sheet, SEC_RATE_ROW, SEC_RATE_COL)

    total_capital_k = dollars_per_kw * system_peak_mw  # $000s/MW × MW = $000s... wait
    # Actually: $/kW × (MW × 1000 kW/MW) / 1000 = $/kW × MW = same number.
    # $12.57/kW × 1078.5 MW = 13,556 ($000s)?  No: $/kW × MW = $/kW × MW.
    # We need $000s: ($/kW × kW) / 1000 = ($/kW × MW × 1000) / 1000 = $/kW × MW.
    # So capital_k = dollars_per_kw × system_peak_mw.  ✓
    total_capital_k = dollars_per_kw * system_peak_mw

    return CostCenterData(
        name="secondary_dist",
        label=CC_LABELS["secondary_dist"],
        n_projects=2,
        total_capacity_mw=0.0,
        total_cost_k=total_capital_k,
        composite_rate=0.0,
        capital_by_year={yr: total_capital_k for yr in YEARS},
        capacity_by_year={yr: system_peak_mw for yr in YEARS},
        is_annual=True,
        is_per_kw_system=True,
    )


# ── MC computation ───────────────────────────────────────────────────────────


def incremental_from_cumulative(by_year: dict[int, float]) -> dict[int, float]:
    """Convert cumulative year-by-year values to year-over-year deltas."""
    result: dict[int, float] = {}
    prev = 0.0
    for yr in YEARS:
        val = by_year.get(yr, 0.0)
        result[yr] = val - prev
        prev = val
    return result


def compute_mc(
    capital_by_year: dict[int, float],
    composite_rate: float,
    escalation: dict[int, float],
    denominator_by_year: dict[int, float],
) -> list[MCRow]:
    """Year-by-year MC.  capital in $000s, denominator in MW ⇒ $/kW."""
    rebase = escalation.get(REBASE_YEAR, 1.0)
    rows: list[MCRow] = []
    for yr in YEARS:
        cap = capital_by_year.get(yr, 0.0)
        esc = escalation.get(yr, 1.0)
        denom = denominator_by_year.get(yr, 0.0)
        rr_nom = cap * composite_rate * esc
        rr_real = cap * composite_rate * rebase
        nom_mc = rr_nom / denom if denom > 0 else 0.0
        real_mc = rr_real / denom if denom > 0 else 0.0
        rows.append(MCRow(yr, cap, rr_nom, esc, nom_mc, real_mc))
    return rows


def levelized(rows: list[MCRow]) -> float:
    """Average of real (base-year) MC over the 2026-2032 levelization window."""
    window = [r for r in rows if r.year in LEVELIZATION_YEARS]
    return sum(r.real_mc for r in window) / len(window) if window else 0.0


# ── CSV export ───────────────────────────────────────────────────────────────


def _sum_mc(
    mc_data: dict[str, list[MCRow]], centers: list[str], yi: int, *, nominal: bool
) -> float:
    return sum(
        (mc_data[c][yi].nominal_mc if nominal else mc_data[c][yi].real_mc)
        for c in centers
    )


def export_levelized_csv(
    mc_data: dict[str, list[MCRow]],
    path: Path,
) -> None:
    bulk_tx_lev = levelized(mc_data["transmission"])
    sub_tx_and_dist_lev = sum(levelized(mc_data[c]) for c in LOCAL_CENTERS)

    rows = [
        {
            "bucket": "bulk_tx",
            "label": CC_LABELS["transmission"],
            "levelized_mc_kw_yr": round(bulk_tx_lev, 2),
            "final_year_real_mc_kw_yr": round(mc_data["transmission"][-1].real_mc, 2),
            "final_year_nominal_mc_kw_yr": round(
                mc_data["transmission"][-1].nominal_mc, 2
            ),
        },
        {
            "bucket": "sub_tx_and_dist",
            "label": CC_LABELS["local_total"],
            "levelized_mc_kw_yr": round(sub_tx_and_dist_lev, 2),
            "final_year_real_mc_kw_yr": round(
                sum(mc_data[c][-1].real_mc for c in LOCAL_CENTERS), 2
            ),
            "final_year_nominal_mc_kw_yr": round(
                sum(mc_data[c][-1].nominal_mc for c in LOCAL_CENTERS), 2
            ),
        },
    ]

    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


def export_annualized_csv(mc_data: dict[str, list[MCRow]], path: Path) -> None:
    rows = []
    for yi, yr in enumerate(YEARS):
        rows.append(
            {
                "year": yr,
                "bulk_tx_nominal": round(mc_data["transmission"][yi].nominal_mc, 2),
                "bulk_tx_real": round(mc_data["transmission"][yi].real_mc, 2),
                "sub_tx_and_dist_nominal": round(
                    _sum_mc(mc_data, LOCAL_CENTERS, yi, nominal=True), 2
                ),
                "sub_tx_and_dist_real": round(
                    _sum_mc(mc_data, LOCAL_CENTERS, yi, nominal=False), 2
                ),
            }
        )
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


# ── Terminal report ──────────────────────────────────────────────────────────


def print_report(
    ccs: dict[str, CostCenterData],
    variants: dict[str, dict[str, list[MCRow]]],
    system_peak_mw: float,
) -> None:
    W = 80
    print("=" * W)
    print("Orange & Rockland MCOS Analysis — Cumulative vs. Incremental")
    print("=" * W)

    print(f"\n── System {'─' * (W - 11)}")
    print(f"  System peak (2024 forecast):  {system_peak_mw:,.1f} MW")

    for name in COST_CENTERS:
        cc = ccs[name]
        if cc.projects:
            print(
                f"\n── {cc.label} — in-service years {'─' * max(1, W - 35 - len(cc.label))}"
            )
            by_year: dict[int, list[ProjectData]] = {}
            for p in cc.projects:
                by_year.setdefault(p.in_service_year, []).append(p)
            for yr in sorted(by_year):
                projs = by_year[yr]
                mw = sum(p.capacity_mw for p in projs)
                cap = sum(p.final_capital_k for p in projs)
                print(
                    f"  {yr}: {len(projs)} project(s), {mw:,.1f} MW, ${cap / 1e3:,.1f}M"
                )

    cum_dil = variants["cumulative_diluted"]
    inc_dil = variants["incremental_diluted"]
    cum_undil = variants["cumulative_undiluted"]
    inc_undil = variants["incremental_undiluted"]

    print(f"\n── Levelized MC ($/kW-yr, real) {'─' * (W - 32)}")
    print(
        f"  {'Cost center':<28} {'Cum.Dil':>8} {'Inc.Dil':>8}"
        f" {'Cum.Und':>8} {'Inc.Und':>8}"
    )
    print(f"  {'─' * (W - 4)}")
    for name in COST_CENTERS:
        cc = ccs[name]
        vals = [
            levelized(cum_dil[name]),
            levelized(inc_dil[name]),
            levelized(cum_undil[name]),
            levelized(inc_undil[name]),
        ]
        parts = "  ".join(f"${v:>6.2f}" for v in vals)
        print(f"  {cc.label:<28} {parts}")
    loc_vals = [
        sum(levelized(v[c]) for c in LOCAL_CENTERS)
        for v in [cum_dil, inc_dil, cum_undil, inc_undil]
    ]
    parts = "  ".join(f"${v:>6.2f}" for v in loc_vals)
    print(f"  {'─' * (W - 4)}")
    print(f"  {'Sub-TX + dist total':<28} {parts}")

    print(f"\n── Year-by-year local_total diluted ($/kW-yr, nominal) {'─' * (W - 55)}")
    print(f"  {'Year':>4}  {'Cumulative':>12}  {'Incremental':>12}  {'Match?':>8}")
    print(f"  {'─' * 42}")
    for yi, yr in enumerate(YEARS):
        cum_v = _sum_mc(cum_dil, LOCAL_CENTERS, yi, nominal=True)
        inc_v = _sum_mc(inc_dil, LOCAL_CENTERS, yi, nominal=True)
        match = "  ✓" if yr == YEARS[0] and abs(cum_v - inc_v) < 0.01 else ""
        print(f"  {yr:>4}  ${cum_v:>10.2f}  ${inc_v:>10.2f}  {match}")

    print()


# ── CLI ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compute MC variants from O&R MCOS workbook"
    )
    parser.add_argument(
        "--path-xlsx",
        type=str,
        required=True,
        help="Path or S3 URL to O&R MCOS workbook",
    )
    parser.add_argument(
        "--system-peak-mw",
        type=float,
        required=True,
        help="Coincident forecast total (MW)",
    )
    parser.add_argument(
        "--path-output-dir", type=Path, required=True, help="Directory for output CSVs"
    )
    args = parser.parse_args()

    print(f"Loading workbook: {args.path_xlsx}")
    wb = _load_wb(args.path_xlsx)

    composite = parse_composite_rates(wb)
    escalation = parse_escalation(wb)
    coinc = parse_coincident_peak(wb)
    print(
        f"  Composite rates: {', '.join(f'{k}={v:.5f}' for k, v in composite.items())}"
    )
    print(f"  Escalation: {escalation[2025]:.4f} → {escalation[2034]:.4f}")
    print(f"  Coincident 2024: {coinc.get(2024, 0):,.1f}, 2025: {coinc[2025]:,.1f} MW")
    print(f"  Using CLI peak: {args.system_peak_mw:,.1f} MW")

    tx_bulk, tx_local = parse_capex_tx_split(wb)
    sub = parse_capex_sub(wb)
    primary = parse_capex_primary(wb)
    sec_dist = parse_capex_secondary(wb, args.system_peak_mw)

    ccs: dict[str, CostCenterData] = {}
    for cc in [tx_bulk, tx_local, sub, primary, sec_dist]:
        cc.composite_rate = composite[cc.name]
        ccs[cc.name] = cc

    print(
        f"  TX bulk: {tx_bulk.n_projects} project (West Nyack), {tx_bulk.total_capacity_mw:,.1f} MW, ${tx_bulk.total_cost_k / 1e3:,.1f}M"
    )
    print(
        f"  TX local: {tx_local.n_projects} projects (Oak St+New Hempstead), {tx_local.total_capacity_mw:,.1f} MW, ${tx_local.total_cost_k / 1e3:,.1f}M"
    )
    print(
        f"  Sub: {sub.n_projects} projects, {sub.total_capacity_mw:,.1f} MW, ${sub.total_cost_k / 1e3:,.1f}M"
    )
    print(
        f"  Primary: {primary.n_projects} projects, {primary.total_capacity_mw:,.1f} MW, ${primary.total_cost_k / 1e3:,.1f}M"
    )
    print(
        f"  SecDist: ${sec_dist.capital_by_year[2025] / args.system_peak_mw:.2f}/kW system (flat)"
    )

    # ── Build all 4 variants ──────────────────────────────────────────────────

    peak_denom = {yr: args.system_peak_mw for yr in YEARS}

    inc_capital: dict[str, dict[int, float]] = {}
    inc_capacity: dict[str, dict[int, float]] = {}
    for name, cc in ccs.items():
        if cc.is_annual:
            inc_capital[name] = cc.capital_by_year
            inc_capacity[name] = cc.capacity_by_year
        else:
            inc_capital[name] = incremental_from_cumulative(cc.capital_by_year)
            inc_capacity[name] = incremental_from_cumulative(cc.capacity_by_year)

    variants: dict[str, dict[str, list[MCRow]]] = {}

    variants["cumulative_diluted"] = {
        name: compute_mc(cc.capital_by_year, cc.composite_rate, escalation, peak_denom)
        for name, cc in ccs.items()
    }

    variants["incremental_diluted"] = {
        name: compute_mc(
            inc_capital[name], ccs[name].composite_rate, escalation, peak_denom
        )
        for name in COST_CENTERS
    }

    variants["cumulative_undiluted"] = {
        name: compute_mc(
            cc.capital_by_year, cc.composite_rate, escalation, cc.capacity_by_year
        )
        for name, cc in ccs.items()
    }

    variants["incremental_undiluted"] = {
        name: compute_mc(
            inc_capital[name],
            ccs[name].composite_rate,
            escalation,
            inc_capacity[name],
        )
        for name in COST_CENTERS
    }

    # ── Export CSVs ───────────────────────────────────────────────────────────

    out = args.path_output_dir
    out.mkdir(parents=True, exist_ok=True)
    print("\nExporting CSVs:")
    for variant_name in VARIANT_NAMES:
        mc_data = variants[variant_name]
        prefix = f"or_{variant_name}"
        export_levelized_csv(mc_data, out / f"{prefix}_levelized.csv")
        export_annualized_csv(mc_data, out / f"{prefix}_annualized.csv")

    print()
    print_report(ccs, variants, args.system_peak_mw)

    wb.close()


if __name__ == "__main__":
    main()
