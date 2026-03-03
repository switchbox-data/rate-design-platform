"""Compute marginal costs from Central Hudson's MCOS study workbook.

Reads CenHud's 2025 MCOS workbook (Demand Side Analytics) and computes
marginal costs by cost center in four variants:

  1. Cumulative diluted    — all in-service projects ÷ system peak
  2. Incremental diluted   — new projects entering in year Y ÷ system peak (BAT input)
  3. Cumulative undiluted  — all in-service projects ÷ in-service project capacity
  4. Incremental undiluted — new projects entering in year Y ÷ new project capacity

Each variant is exported in both annualized (year-by-year) and levelized form,
for a total of 8 CSVs.

CenHud's workbook provides three cost centers:
  1. Local Transmission (69/115 kV) — all local, no FERC-jurisdictional bulk TX
  2. Substation — distribution substations
  3. Feeder Circuit — primary distribution feeders

All three are INCLUDED in the BAT input (no exclusion needed).

MC formula (diluted):
  Real MC(Y)    = sum[ annual_cost_per_kW(p) × capacity_kW(p) ] / system_peak_kW
  Nominal MC(Y) = Real MC(Y) × escalation(Y)

  The workbook uses a peak-share formula instead (cost_per_kW × peak_share),
  but we use capacity-based for cross-utility consistency with ConEd/O&R/NiMo.

MC formula (undiluted):
  Real MC(Y)    = sum[ annual_cost(p) × cap(p) ] / sum[ cap(p) ]
  Nominal MC(Y) = Real MC(Y) × escalation(Y)

CenHud's workbook provides FLAT NOMINAL costs (no escalation), but we apply
a 2.1%/yr GDP deflator (base year 2026) for consistency with ConEd/O&R/NiMo.
Real MC uses the workbook's flat values; nominal MC is escalated.

Costs contribute starting the year AFTER the project's in-service year.

Outputs (8 CSVs):
  - cenhud_{cumulative,incremental}_{diluted,undiluted}_{levelized,annualized}.csv
  - Terminal report (always printed)

Usage (via Justfile):
    just analyze-cenhud
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import fsspec
import openpyxl
import polars as pl


# ── Study parameters ─────────────────────────────────────────────────────────

YEARS = list(range(2026, 2036))  # 2026 through 2035
N_YEARS = len(YEARS)
BASE_YEAR = 2026
GDP_DEFLATOR_RATE = 0.021  # 2.1%/yr, matching NiMo / ConEd / O&R steady-state rate

COST_CENTERS = ["local_tx", "substation", "feeder"]
BUCKET_KEYS = [*COST_CENTERS, "total"]

BUCKET_LABELS = {
    "local_tx": "Local Transmission (69/115 kV)",
    "substation": "Substation",
    "feeder": "Feeder Circuit",
    "total": "Total (all cost centers)",
}

VARIANT_NAMES = [
    "cumulative_diluted",
    "incremental_diluted",
    "cumulative_undiluted",
    "incremental_undiluted",
]

# Per-sheet project column positions (1-indexed openpyxl columns).
# Derived from workbook exploration: projects alternate with spacer columns.
SHEET_CONFIG: dict[str, dict] = {
    "local_tx": {
        "sheet_name": "Local Transmission",
        "project_cols": {
            "Future Unidentified": 5,
            "Northwest 115/69": 7,
            "RD-RJ Lines": 8,
        },
    },
    "substation": {
        "sheet_name": "Substation",
        "project_cols": {
            "Future Unidentified": 5,
            "Maybrook": 7,
            "Pulvers 13kV": 8,
            "Woodstock": 9,
            "New Baltimore": 10,
            "Hurley Ave": 11,
        },
    },
    "feeder": {
        "sheet_name": "Feeder",
        "project_cols": {
            "Future Unidentified": 5,
            "WI_8031": 7,
        },
    },
}


# ── Data structures ──────────────────────────────────────────────────────────


@dataclass
class McosProject:
    """One project from the MCOS workbook."""

    name: str
    cost_center: str
    in_service_year: int
    capacity_kw: float
    annual_cost_per_kw: float  # $/kW-yr (row 26, flat nominal)
    peak_share: float  # fraction of CH coincident peak


@dataclass
class MCRow:
    """One year's MC for a bucket (used for all variants)."""

    year: int
    cost_real_k: float  # numerator in $000s
    cost_nominal_k: float  # same (flat nominal)
    denominator_mw: float
    nominal_mc: float  # $/kW-yr
    real_mc: float  # $/kW-yr (same as nominal — flat costs)


# ── Workbook parsing ─────────────────────────────────────────────────────────


def _find_row_by_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: int) -> int:
    """Find row where column A (col 1) has the given integer label."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == label:
            return r
    msg = f"Row with label {label} not found in {ws.title!r}"
    raise ValueError(msg)


def _find_row_by_text(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    text: str,
    *,
    search_cols: tuple[int, ...] = (2, 3),
) -> int:
    """Find first row where any of `search_cols` contains `text` as substring."""
    for r in range(1, ws.max_row + 1):
        for c in search_cols:
            val = ws.cell(r, c).value
            if isinstance(val, str) and text in val:
                return r
    msg = f"Row containing {text!r} not found in {ws.title!r}"
    raise ValueError(msg)


def parse_projects(wb: openpyxl.Workbook) -> list[McosProject]:
    """Extract per-project data from the three MCOS calculation sheets."""
    projects: list[McosProject] = []

    for cc_key, cfg in SHEET_CONFIG.items():
        ws = wb[cfg["sheet_name"]]
        insvc_row = _find_row_by_text(ws, "In Service Year")
        cap_row = _find_row_by_label(ws, -3)
        cost_row = _find_row_by_label(ws, -26)
        share_row = _find_row_by_text(ws, "Share of Central Hudson")

        for name, col in cfg["project_cols"].items():
            insvc_val = ws.cell(insvc_row, col).value
            in_service_year = int(insvc_val)

            projects.append(
                McosProject(
                    name=name,
                    cost_center=cc_key,
                    in_service_year=in_service_year,
                    capacity_kw=float(ws.cell(cap_row, col).value),
                    annual_cost_per_kw=float(ws.cell(cost_row, col).value),
                    peak_share=float(ws.cell(share_row, col).value),
                )
            )

    return projects


# ── Validation data from Summary sheet ───────────────────────────────────────


def parse_summary_tables(
    wb: openpyxl.Workbook,
) -> tuple[dict[str, list[float]], dict[str, list[float]]]:
    """Read Tables 1 (undiluted) and 2 (diluted) from Summary - System sheet.

    Returns (diluted, undiluted) dicts keyed by cost center with lists of
    10 annual values.
    """
    ws = wb["Summary - System"]

    # Table 2 (system-wide / diluted): rows 6-15, cols B(2) C(3) D(4) E(5)
    diluted: dict[str, list[float]] = {
        "local_tx": [float(ws.cell(r, 2).value or 0) for r in range(6, 16)],
        "substation": [float(ws.cell(r, 3).value or 0) for r in range(6, 16)],
        "feeder": [float(ws.cell(r, 4).value or 0) for r in range(6, 16)],
        "total": [float(ws.cell(r, 5).value or 0) for r in range(6, 16)],
    }

    # Table 1 (areas-with-projects / undiluted): rows 21-30, cols B(2) C(3) D(4)
    undiluted: dict[str, list[float]] = {
        "local_tx": [float(ws.cell(r, 2).value or 0) for r in range(21, 31)],
        "substation": [float(ws.cell(r, 3).value or 0) for r in range(21, 31)],
        "feeder": [float(ws.cell(r, 4).value or 0) for r in range(21, 31)],
    }

    return diluted, undiluted


# ── MC computation ───────────────────────────────────────────────────────────


def _projects_by_bucket(
    projects: list[McosProject],
) -> dict[str, list[McosProject]]:
    """Group projects by cost center and create the 'total' bucket."""
    by_cc: dict[str, list[McosProject]] = {cc: [] for cc in COST_CENTERS}
    for p in projects:
        by_cc[p.cost_center].append(p)
    by_cc["total"] = list(projects)
    return by_cc


def _escalation(year: int) -> float:
    """GDP deflator escalation factor for a given year (base year = 1.0)."""
    return (1 + GDP_DEFLATOR_RATE) ** (year - BASE_YEAR)


def compute_bucket_mc(
    projects: list[McosProject],
    system_peak_mw: float,
    *,
    cumulative: bool,
    diluted: bool,
) -> list[MCRow]:
    """Year-by-year MC for a set of projects under one variant.

    cumulative=True:  in-scope = projects with in_service_year + 1 <= year
    cumulative=False: in-scope = projects with in_service_year + 1 == year
    diluted=True:     denominator = system peak
    diluted=False:    denominator = sum(capacity_kw) of in-scope projects

    The workbook's costs are flat nominal (no escalation).  We treat them as
    base-year (2026) real values and apply a 2.1%/yr GDP deflator to produce
    nominal values, consistent with ConEd/O&R/NiMo.
    """
    system_peak_kw = system_peak_mw * 1000
    rows: list[MCRow] = []

    for year in YEARS:
        esc = _escalation(year)

        if cumulative:
            in_scope = [p for p in projects if p.in_service_year + 1 <= year]
        else:
            in_scope = [p for p in projects if p.in_service_year + 1 == year]

        if not in_scope:
            rows.append(
                MCRow(
                    year=year,
                    cost_real_k=0.0,
                    cost_nominal_k=0.0,
                    denominator_mw=system_peak_mw if diluted else 0.0,
                    nominal_mc=0.0,
                    real_mc=0.0,
                )
            )
            continue

        if diluted:
            cost_real = sum(p.annual_cost_per_kw * p.capacity_kw for p in in_scope)
            denom_kw = system_peak_kw
            real_mc = cost_real / denom_kw
        else:
            cost_real = sum(p.annual_cost_per_kw * p.capacity_kw for p in in_scope)
            denom_kw = sum(p.capacity_kw for p in in_scope)
            real_mc = cost_real / denom_kw if denom_kw > 0 else 0.0

        nominal_mc = real_mc * esc

        rows.append(
            MCRow(
                year=year,
                cost_real_k=cost_real / 1000,
                cost_nominal_k=cost_real * esc / 1000,
                denominator_mw=denom_kw / 1000,
                nominal_mc=nominal_mc,
                real_mc=real_mc,
            )
        )

    return rows


def levelized(rows: list[MCRow]) -> float:
    """Average of real (base-year) MC across all study years."""
    return sum(r.real_mc for r in rows) / len(rows) if rows else 0.0


# ── CSV export ───────────────────────────────────────────────────────────────


def export_levelized_csv(
    mc_data: dict[str, list[MCRow]],
    bucket_projects: dict[str, list[McosProject]],
    path: Path,
) -> None:
    rows = []
    for key in BUCKET_KEYS:
        mc_rows = mc_data[key]
        projs = bucket_projects[key]
        lev = levelized(mc_rows)
        final_real = mc_rows[-1].real_mc
        final_nom = mc_rows[-1].nominal_mc
        rows.append(
            {
                "bucket": key,
                "label": BUCKET_LABELS[key],
                "n_projects": len(projs),
                "capacity_mw": round(sum(p.capacity_kw for p in projs) / 1000, 1),
                "levelized_mc_kw_yr": round(lev, 2),
                "final_year_real_mc_kw_yr": round(final_real, 2),
                "final_year_nominal_mc_kw_yr": round(final_nom, 2),
            }
        )
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


def export_annualized_csv(mc_data: dict[str, list[MCRow]], path: Path) -> None:
    rows = []
    for yi, yr in enumerate(YEARS):
        row: dict[str, object] = {"year": yr}
        for key in BUCKET_KEYS:
            row[f"{key}_nominal"] = round(mc_data[key][yi].nominal_mc, 2)
            row[f"{key}_real"] = round(mc_data[key][yi].real_mc, 2)
        rows.append(row)
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


# ── Terminal report ──────────────────────────────────────────────────────────


def print_report(
    projects: list[McosProject],
    variants: dict[str, dict[str, list[MCRow]]],
    system_peak_mw: float,
    diluted_table: dict[str, list[float]],
    undiluted_table: dict[str, list[float]],
) -> None:
    W = 80
    print("=" * W)
    print("Central Hudson MCOS Analysis — Cumulative vs. Incremental")
    print("=" * W)

    print(f"\n── System {'─' * (W - 11)}")
    print(f"  System peak (2024 actual):  {system_peak_mw:,.0f} MW")
    print(f"  Study period:               {YEARS[0]}–{YEARS[-1]} ({N_YEARS} years)")
    print("  Workbook escalation:        None (flat nominal)")
    print(
        f"  Applied escalation:         {GDP_DEFLATOR_RATE * 100:.1f}%/yr GDP deflator (base {BASE_YEAR})"
    )
    print(f"  Total projects:             {len(projects)}")

    # Per-cost-center summary
    print(f"\n── Projects {'─' * (W - 12)}")
    for cc in COST_CENTERS:
        cc_projs = [p for p in projects if p.cost_center == cc]
        total_cap = sum(p.capacity_kw for p in cc_projs) / 1000
        print(f"\n  {BUCKET_LABELS[cc]}:")
        for p in cc_projs:
            flag = " (future unidentified)" if "Future" in p.name else ""
            print(
                f"    {p.name:25s}  in-svc={p.in_service_year}  "
                f"cap={p.capacity_kw / 1000:8.1f} MW  "
                f"cost=${p.annual_cost_per_kw:7.2f}/kW-yr  "
                f"share={p.peak_share * 100:5.2f}%{flag}"
            )
        print(f"    {'total':25s}  {' ' * 15}cap={total_cap:8.1f} MW")

    # Levelized summary for all 4 variants
    print(f"\n── Levelized MC ($/kW-yr) {'─' * (W - 26)}")
    for vname in VARIANT_NAMES:
        mc_data = variants[vname]
        print(f"\n  {vname}:")
        for key in BUCKET_KEYS:
            lev = levelized(mc_data[key])
            print(f"    {BUCKET_LABELS[key]:40s}  ${lev:7.2f}/kW-yr")

    # Comparison against workbook Tables 1 and 2
    print(f"\n── Comparison vs. workbook tables {'─' * (W - 35)}")
    cum_dil = variants["cumulative_diluted"]

    print("\n  Table 2 (diluted) — workbook uses peak-share; we use capacity-based.")
    print("  Values differ by design (see README §7A).\n")
    print(f"  {'':15s}  {'Ours (capacity)':>16s}  {'Workbook (peak-share)':>22s}")
    for cc in COST_CENTERS:
        if cc not in diluted_table:
            continue
        ours_lev = levelized(cum_dil[cc])
        wb_lev = sum(diluted_table[cc]) / len(diluted_table[cc])
        print(
            f"  {BUCKET_LABELS[cc]:40s}  ${ours_lev:7.2f}/kW-yr  ${wb_lev:7.2f}/kW-yr"
        )
    if "total" in diluted_table:
        ours_total = levelized(cum_dil["total"])
        wb_total = sum(diluted_table["total"]) / len(diluted_table["total"])
        print(
            f"  {BUCKET_LABELS['total']:40s}  ${ours_total:7.2f}/kW-yr"
            f"  ${wb_total:7.2f}/kW-yr"
        )

    print("\n  Note: Workbook Table 1 ('areas with projects') uses peak-share-based")
    print("  aggregation; our undiluted variant uses project capacity. Values differ")
    print("  by design — see README for details.")

    # First-year sanity check: cumulative == incremental in the first contributing year
    print(f"\n── First-year sanity check {'─' * (W - 28)}")
    inc_dil = variants["incremental_diluted"]
    for cc in COST_CENTERS:
        for yi in range(N_YEARS):
            if cum_dil[cc][yi].nominal_mc > 0:
                cum_val = cum_dil[cc][yi].nominal_mc
                inc_val = inc_dil[cc][yi].nominal_mc
                match = abs(cum_val - inc_val) < 0.001
                print(
                    f"  {BUCKET_LABELS[cc]:40s}  year={YEARS[yi]}  "
                    f"cum=${cum_val:.4f}  inc=${inc_val:.4f}  "
                    f"{'✓' if match else '✗'}"
                )
                break

    print("\n" + "=" * W)


# ── Main ─────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(description="CenHud MCOS analysis")
    parser.add_argument("--path-xlsx", required=True)
    parser.add_argument("--system-peak-mw", type=float, required=True)
    parser.add_argument("--path-output-dir", required=True)
    args = parser.parse_args()

    out_dir = Path(args.path_output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    system_peak_mw = args.system_peak_mw

    # Load workbook
    print(f"Loading workbook from {args.path_xlsx} ...")
    with fsspec.open(args.path_xlsx, "rb") as f:
        wb = openpyxl.load_workbook(f, data_only=True)

    projects = parse_projects(wb)
    print(f"Parsed {len(projects)} projects across {len(COST_CENTERS)} cost centers")

    bucket_projects = _projects_by_bucket(projects)

    # Compute all 4 variants
    variants: dict[str, dict[str, list[MCRow]]] = {}
    for vname in VARIANT_NAMES:
        cumulative = "cumulative" in vname
        diluted = "undiluted" not in vname
        mc_data: dict[str, list[MCRow]] = {}
        for key in BUCKET_KEYS:
            mc_data[key] = compute_bucket_mc(
                bucket_projects[key],
                system_peak_mw,
                cumulative=cumulative,
                diluted=diluted,
            )
        variants[vname] = mc_data

    # Export CSVs
    print("\nExporting CSVs:")
    for vname in VARIANT_NAMES:
        export_levelized_csv(
            variants[vname],
            bucket_projects,
            out_dir / f"cenhud_{vname}_levelized.csv",
        )
        export_annualized_csv(
            variants[vname],
            out_dir / f"cenhud_{vname}_annualized.csv",
        )

    # Validation data
    diluted_table, undiluted_table = parse_summary_tables(wb)

    # Terminal report
    print_report(projects, variants, system_peak_mw, diluted_table, undiluted_table)


if __name__ == "__main__":
    main()
