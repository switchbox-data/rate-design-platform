"""Compute marginal costs from a CRA-prepared NYSEG MCOS workbook.

Reads system-wide diluted MC from T1A, undiluted from T2, and capacity
data from T7.  Independently derives division-peak-weighted MC from
T4/T4B/T5/T6 for verification against the workbook's Table 2.

Cost centers (at primary voltage level):
  1. Upstream — substation (115 kV/46 kV) + feeder (115 kV/34.5 kV)
  2. Distribution Substation (12.5 kV)
  3. Primary Feeder (12.5 kV/4 kV)

Four MC variants (8 CSVs):
  1. Cumulative diluted    — accumulated costs ÷ system peak
  2. Incremental diluted   — new-year costs ÷ system peak (BAT input)
  3. Cumulative undiluted  — accumulated costs ÷ accumulated capacity
  4. Incremental undiluted — new-year costs ÷ new-year capacity

CRA applies 2.0%/yr inflation to project costs from in-service year.
The diluted denominator is the 2035 forecast peak (fixed across years).
Cumulative variants inflate prior years' incremental contributions at 2%/yr.

The "Total at Primary" column includes demand-related loss factors to
express the combined MC at primary distribution voltage; individual cost
center columns are at their own voltage level (upstream, dist sub, primary).

Data sources within the workbook:
  T1A  — Table 3/7 (diluted year-by-year, after within-division adjustment)
       — Table 2 (division-peak-weighted year-by-year, for verification)
  T2   — Undiluted year-by-year (capacity-weighted per sub-type)
  T4   — Division-level upstream sub MC  (verification)
  T4B  — Division-level upstream feeder MC (verification)
  T5   — Division-level dist sub MC (NYSEG only; verification)
  T6   — Division-level primary feeder MC (NYSEG only; verification)
  T7   — Per-division per-year investment + capacity

Usage (via Justfile):
    just analyze-nyseg
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import fsspec
import openpyxl
import polars as pl


# ── Study parameters ─────────────────────────────────────────────────────────

YEARS = list(range(2026, 2036))
N_YEARS = len(YEARS)
INFLATION_RATE = 0.02
WACC = 0.06975

COST_CENTERS = ["upstream", "dist_sub", "primary_feeder"]
BUCKET_KEYS = [*COST_CENTERS, "total"]
BUCKET_LABELS = {
    "upstream": "Upstream (Sub + Feeder)",
    "dist_sub": "Distribution Substation",
    "primary_feeder": "Primary Feeder",
    "total": "Total at Primary",
}

VARIANT_NAMES = [
    "cumulative_diluted",
    "incremental_diluted",
    "cumulative_undiluted",
    "incremental_undiluted",
]

# T7 capacity-section column positions (1-indexed openpyxl)
T7_CAP_YEAR = 10
T7_CAP_UPS_SUB = 11
T7_CAP_UPS_FEED = 12
T7_CAP_DIST_SUB = 13
T7_CAP_PF = 14


# ── Configuration ────────────────────────────────────────────────────────────


@dataclass
class CRAConfig:
    """Configuration for a CRA-style NYSEG/RGE MCOS workbook."""

    utility: str
    display_name: str
    divisions: list[str]
    system_peak_mw: float
    t4_pattern: str  # sheet-name substring for upstream sub
    t4b_pattern: str  # sheet-name substring for upstream feeder
    div_col_start: int = 3
    has_division_t5_t6: bool = True  # RGE T5/T6 have different layout


NYSEG_DIVISIONS = [
    "Auburn",
    "Binghamton",
    "Brewster",
    "Elmira",
    "Geneva",
    "Hornell",
    "Ithaca",
    "Lancaster",
    "Liberty",
    "Lockport",
    "Mechanicville",
    "Oneonta",
    "Plattsburgh",
]


# ── Workbook navigation ─────────────────────────────────────────────────────


def find_sheet(wb: openpyxl.Workbook, pattern: str) -> Any:
    """Find first sheet whose name contains ``pattern``."""
    for name in wb.sheetnames:
        if pattern in name:
            return wb[name]
    msg = f"No sheet matching '{pattern}' in {wb.sheetnames}"
    raise ValueError(msg)


def find_year_blocks(ws: Any, year_col: int = 2) -> list[int]:
    """Return row numbers where ``year_col`` == 2026 (start of a year block)."""
    blocks: list[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, year_col).value
        if v is None:
            continue
        try:
            if int(float(v)) == YEARS[0]:
                blocks.append(r)
        except (ValueError, TypeError):
            pass
    return blocks


def read_year_block(
    ws: Any,
    start_row: int,
    cols: dict[str, int],
    year_col: int = 2,
) -> dict[str, list[float]]:
    """Read 10 contiguous year-rows starting at ``start_row``."""
    result: dict[str, list[float]] = {k: [] for k in cols}
    for offset in range(N_YEARS):
        r = start_row + offset
        yr_val = ws.cell(r, year_col).value
        if yr_val is None or int(float(yr_val)) != YEARS[offset]:
            msg = f"Expected year {YEARS[offset]} at row {r}, got {yr_val!r}"
            raise ValueError(msg)
        for key, col in cols.items():
            v = ws.cell(r, col).value
            result[key].append(float(v) if v is not None else 0.0)
    return result


# ── Parsing: T1A (diluted) ───────────────────────────────────────────────────


def parse_diluted_incremental(
    wb: openpyxl.Workbook,
) -> tuple[dict[str, list[float]], dict[str, list[float]], float]:
    """Read incremental diluted MC and division-weighted Table 2 from T1A.

    Returns ``(diluted_inc, table2_wb, diluted_levelized_total)``.
    """
    ws = find_sheet(wb, "T1A")
    blocks = find_year_blocks(ws)
    if len(blocks) < 2:
        msg = f"Expected ≥2 year blocks in T1A, found {len(blocks)}"
        raise ValueError(msg)

    cols = {"upstream": 3, "dist_sub": 4, "primary_feeder": 5, "total": 6}
    table2 = read_year_block(ws, blocks[0], cols)
    diluted = read_year_block(ws, blocks[1], cols)

    lev_total = 0.0
    for r in range(blocks[1] + N_YEARS, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and "levelized" in v.lower():
            lev_total = float(ws.cell(r, 6).value or 0)
            break

    return diluted, table2, lev_total


# ── Parsing: T2 (undiluted) ──────────────────────────────────────────────────


def parse_undiluted_incremental(
    wb: openpyxl.Workbook,
) -> tuple[dict[str, list[float]], float]:
    """Read incremental undiluted MC per sub-type from T2.

    Returns ``(per_subtype_data, undiluted_levelized_total)``.
    Keys: ``ups_sub``, ``ups_feed``, ``dist_sub``, ``primary_feeder``, ``total``.
    """
    ws = find_sheet(wb, "T2")
    blocks = find_year_blocks(ws)
    if not blocks:
        msg = "No year block found in T2"
        raise ValueError(msg)

    cols = {
        "ups_sub": 3,
        "ups_feed": 4,
        "dist_sub": 5,
        "primary_feeder": 6,
        "total": 7,
    }
    data = read_year_block(ws, blocks[0], cols)

    lev_total = 0.0
    for r in range(blocks[0] + N_YEARS, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and "levelized" in v.lower():
            lev_total = float(ws.cell(r, 7).value or 0)
            break

    return data, lev_total


# ── Parsing: T7 (capacity) ──────────────────────────────────────────────────


def parse_t7_capacity(wb: openpyxl.Workbook) -> dict[int, dict[str, float]]:
    """Sum per-sub-type capacity (MVA) across all divisions from T7."""
    ws = find_sheet(wb, "T7")
    cap_cols = {
        "ups_sub": T7_CAP_UPS_SUB,
        "ups_feed": T7_CAP_UPS_FEED,
        "dist_sub": T7_CAP_DIST_SUB,
        "pf": T7_CAP_PF,
    }
    result: dict[int, dict[str, float]] = {y: {k: 0.0 for k in cap_cols} for y in YEARS}

    for r in range(17, ws.max_row + 1):
        yr_val = ws.cell(r, T7_CAP_YEAR).value
        if yr_val is None:
            continue
        try:
            yr = int(float(yr_val))
        except (ValueError, TypeError):
            continue
        if yr not in result:
            continue
        for key, col in cap_cols.items():
            v = ws.cell(r, col).value
            if v is not None:
                try:
                    result[yr][key] += float(v)
                except (ValueError, TypeError):
                    pass

    return result


# ── Parsing: T4/T4B/T5/T6 (division-level, for verification) ────────────────


def parse_division_mc_and_peaks(
    wb: openpyxl.Workbook,
    config: CRAConfig,
) -> tuple[dict[str, dict[str, list[float]]], dict[str, float]]:
    """Read per-division MC from T4/T4B (and T5/T6 if available) and peak loads.

    Returns ``(div_mc, peaks)``.
    """
    div_mc: dict[str, dict[str, list[float]]] = {}
    n_div = len(config.divisions)

    ws_t4 = find_sheet(wb, config.t4_pattern)
    div_mc["upstream_sub"] = _read_division_block(
        ws_t4, config.divisions, config.div_col_start, n_div
    )

    ws_t4b = find_sheet(wb, config.t4b_pattern)
    div_mc["upstream_feed"] = _read_division_block(
        ws_t4b, config.divisions, config.div_col_start, n_div
    )

    if config.has_division_t5_t6:
        ws_t5 = find_sheet(wb, "T5")
        div_mc["dist_sub"] = _read_division_block(
            ws_t5, config.divisions, config.div_col_start, n_div
        )
        ws_t6 = find_sheet(wb, "T6")
        div_mc["primary_feeder"] = _read_division_block(
            ws_t6, config.divisions, config.div_col_start, n_div
        )

    peaks: dict[str, float] = {}
    for i, name in enumerate(config.divisions):
        col = config.div_col_start + i
        v = ws_t4.cell(20, col).value
        peaks[name] = float(v) if v is not None else 0.0

    return div_mc, peaks


def _read_division_block(
    ws: Any,
    divisions: list[str],
    col_start: int,
    n_div: int,
) -> dict[str, list[float]]:
    """Read MC for each division from rows 7–16 of a division sheet."""
    result: dict[str, list[float]] = {}
    for i in range(n_div):
        col = col_start + i
        values = [
            float(ws.cell(r, col).value) if ws.cell(r, col).value is not None else 0.0
            for r in range(7, 17)
        ]
        result[divisions[i]] = values
    return result


# ── Computation ──────────────────────────────────────────────────────────────


def combine_upstream_undiluted(
    ups_sub_mc: list[float],
    ups_feed_mc: list[float],
    cap_by_year: dict[int, dict[str, float]],
) -> list[float]:
    """Capacity-weighted combination of upstream sub + feeder undiluted MC."""
    result: list[float] = []
    for yi, yr in enumerate(YEARS):
        cap_s = cap_by_year[yr]["ups_sub"]
        cap_f = cap_by_year[yr]["ups_feed"]
        total_cap = cap_s + cap_f
        if total_cap > 0:
            mc = (ups_sub_mc[yi] * cap_s + ups_feed_mc[yi] * cap_f) / total_cap
        else:
            mc = 0.0
        result.append(mc)
    return result


def accumulate_cumulative_diluted(
    incremental: dict[str, list[float]],
) -> dict[str, list[float]]:
    """Accumulate incremental diluted into cumulative with 2%/yr inflation.

    ``cumulative(Y) = Σ_{t≤Y} incremental(t) × (1.02)^(Y−t)``
    """
    result: dict[str, list[float]] = {}
    for key in BUCKET_KEYS:
        inc = incremental[key]
        cum: list[float] = []
        for yi in range(N_YEARS):
            total = sum(
                inc[ti] * (1 + INFLATION_RATE) ** (yi - ti) for ti in range(yi + 1)
            )
            cum.append(total)
        result[key] = cum
    return result


def accumulate_cumulative_undiluted(
    inc_mc: dict[str, list[float]],
    cap_by_year: dict[int, dict[str, float]],
) -> dict[str, list[float]]:
    """Accumulate undiluted incremental into cumulative.

    ``cumulative(Y) = Σ_{t≤Y}[mc(t)×cap(t)×(1.02)^(Y−t)] / Σ_{t≤Y}[cap(t)]``
    """
    cap_keys: dict[str, list[str]] = {
        "upstream": ["ups_sub", "ups_feed"],
        "dist_sub": ["dist_sub"],
        "primary_feeder": ["pf"],
        "total": ["ups_sub", "ups_feed", "dist_sub", "pf"],
    }

    result: dict[str, list[float]] = {}
    for bucket in BUCKET_KEYS:
        cum: list[float] = []
        for yi in range(N_YEARS):
            num = 0.0
            den = 0.0
            for ti in range(yi + 1):
                yr_t = YEARS[ti]
                cap_t = sum(cap_by_year[yr_t][k] for k in cap_keys[bucket])
                cost_t = inc_mc[bucket][ti] * cap_t
                num += cost_t * (1 + INFLATION_RATE) ** (yi - ti)
                den += cap_t
            cum.append(num / den if den > 0 else 0.0)
        result[bucket] = cum
    return result


def derive_table2(
    div_mc: dict[str, dict[str, list[float]]],
    peaks: dict[str, float],
    system_peak: float,
    has_t5_t6: bool,
) -> dict[str, list[float]]:
    """Derive Table 2 (division-peak-weighted MC) from division sheets."""
    derived: dict[str, list[float]] = {}

    ups_sub = div_mc.get("upstream_sub", {})
    ups_feed = div_mc.get("upstream_feed", {})
    upstream: list[float] = []
    for yi in range(N_YEARS):
        total = sum(
            (ups_sub.get(d, [0.0] * N_YEARS)[yi] + ups_feed.get(d, [0.0] * N_YEARS)[yi])
            * peaks.get(d, 0.0)
            for d in peaks
        )
        upstream.append(total / system_peak if system_peak > 0 else 0.0)
    derived["upstream"] = upstream

    if has_t5_t6:
        for cc in ["dist_sub", "primary_feeder"]:
            mc_by_div = div_mc.get(cc, {})
            annual: list[float] = []
            for yi in range(N_YEARS):
                total = sum(
                    mc_by_div.get(d, [0.0] * N_YEARS)[yi] * peaks.get(d, 0.0)
                    for d in peaks
                )
                annual.append(total / system_peak if system_peak > 0 else 0.0)
            derived[cc] = annual

    if has_t5_t6:
        derived["total"] = [
            sum(derived[cc][yi] for cc in COST_CENTERS) for yi in range(N_YEARS)
        ]
    else:
        derived["total"] = upstream[:]

    return derived


def levelized_npv(values: list[float], wacc: float = WACC) -> float:
    """NPV-based levelized cost (matching CRA workbook convention)."""
    pv = sum(v / (1 + wacc) ** i for i, v in enumerate(values))
    annuity = sum(1 / (1 + wacc) ** i for i in range(len(values)))
    return pv / annuity if annuity > 0 else 0.0


def to_real(nominal: list[float]) -> list[float]:
    """Convert nominal values to base-year (2026) real dollars."""
    return [v / (1 + INFLATION_RATE) ** i for i, v in enumerate(nominal)]


# ── CSV export ───────────────────────────────────────────────────────────────


def export_annualized_csv(mc_nominal: dict[str, list[float]], path: Path) -> None:
    rows: list[dict[str, object]] = []
    for yi, yr in enumerate(YEARS):
        row: dict[str, object] = {"year": yr}
        for key in BUCKET_KEYS:
            nom = mc_nominal[key][yi]
            real_val = nom / (1 + INFLATION_RATE) ** yi
            row[f"{key}_nominal"] = round(nom, 4)
            row[f"{key}_real"] = round(real_val, 4)
        rows.append(row)
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


def export_levelized_csv(mc_nominal: dict[str, list[float]], path: Path) -> None:
    rows: list[dict[str, object]] = []
    for key in BUCKET_KEYS:
        values = mc_nominal[key]
        real_values = to_real(values)
        lev_real = sum(real_values) / len(real_values) if real_values else 0.0
        lev_npv = levelized_npv(values)
        rows.append(
            {
                "bucket": key,
                "label": BUCKET_LABELS[key],
                "levelized_mc_kw_yr": round(lev_real, 4),
                "levelized_npv_mc_kw_yr": round(lev_npv, 4),
            }
        )
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


# ── Terminal report ──────────────────────────────────────────────────────────


def print_report(
    config: CRAConfig,
    variants: dict[str, dict[str, list[float]]],
    table2_derived: dict[str, list[float]],
    table2_wb: dict[str, list[float]],
    diluted_lev_wb: float,
    undiluted_lev_wb: float,
) -> None:
    W = 80
    print("=" * W)
    print(f"{config.display_name} MCOS Analysis — CRA Methodology")
    print("=" * W)

    print(f"\n── System {'─' * (W - 11)}")
    print(f"  System peak (2035 forecast):  {config.system_peak_mw:,.2f} MW")
    print(f"  Study period:                 {YEARS[0]}–{YEARS[-1]} ({N_YEARS} years)")
    print(f"  Inflation:                    {INFLATION_RATE * 100:.1f}%/yr")
    print(f"  WACC:                         {WACC * 100:.3f}%")
    print(f"  Divisions:                    {len(config.divisions)}")

    print(f"\n── Levelized MC (real $/kW-yr, simple avg) {'─' * max(1, W - 43)}")
    header = f"  {'Variant':<28}"
    for k in BUCKET_KEYS:
        header += f" {BUCKET_LABELS[k]:>14}"
    print(header)
    print(f"  {'─' * (W - 4)}")
    for vname in VARIANT_NAMES:
        data = variants[vname]
        real_values = {k: to_real(data[k]) for k in BUCKET_KEYS}
        levs = {k: sum(real_values[k]) / N_YEARS for k in BUCKET_KEYS}
        line = f"  {vname:<28}"
        for k in BUCKET_KEYS:
            line += f" ${levs[k]:>12.2f}"
        print(line)

    # Table 2 verification
    available = [k for k in BUCKET_KEYS if k in table2_derived and k in table2_wb]
    if available:
        print(f"\n── Table 2 verification {'─' * (W - 24)}")
        for cc in available:
            max_delta = max(
                abs(table2_derived[cc][yi] - table2_wb[cc][yi]) for yi in range(N_YEARS)
            )
            print(f"  {BUCKET_LABELS.get(cc, cc):30s}  max |Δ| = {max_delta:.6f}")

    # Levelized validation
    print(f"\n── Levelized validation (NPV-based) {'─' * max(1, W - 37)}")
    our_dil = levelized_npv(variants["incremental_diluted"]["total"])
    print(
        f"  Diluted total at primary:   ours=${our_dil:.4f}  wb=${diluted_lev_wb:.4f}"
    )
    our_und = levelized_npv(variants["incremental_undiluted"]["total"])
    print(
        f"  Undiluted total at primary: ours=${our_und:.4f}  wb=${undiluted_lev_wb:.4f}"
    )

    # Year-by-year incremental diluted
    print(
        f"\n── Year-by-year incremental diluted (nominal $/kW-yr) "
        f"{'─' * max(1, W - 52)}"
    )
    inc = variants["incremental_diluted"]
    header2 = f"  {'Year':>4}"
    for k in BUCKET_KEYS:
        header2 += f" {k:>12}"
    print(header2)
    for yi in range(N_YEARS):
        line = f"  {YEARS[yi]:>4}"
        for k in BUCKET_KEYS:
            line += f" ${inc[k][yi]:>10.2f}"
        print(line)

    print("\n" + "=" * W)


# ── Pipeline ─────────────────────────────────────────────────────────────────


def run_analysis(
    xlsx_path: str,
    config: CRAConfig,
    output_dir: Path,
) -> None:
    """Run the full CRA MCOS analysis pipeline."""
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading {config.display_name} workbook from {xlsx_path} ...")
    if xlsx_path.startswith("s3://"):
        with fsspec.open(xlsx_path, "rb") as f:
            wb = openpyxl.load_workbook(f, data_only=True)
    else:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1. Diluted from T1A
    diluted_inc, table2_wb, diluted_lev_wb = parse_diluted_incremental(wb)
    print("  T1A: incremental diluted + Table 2 parsed")

    # 2. Undiluted from T2
    undiluted_raw, undiluted_lev_wb = parse_undiluted_incremental(wb)
    print("  T2: incremental undiluted per sub-type parsed")

    # 3. Capacity from T7
    cap = parse_t7_capacity(wb)
    print("  T7: capacity by year parsed")

    # 4. Division data for verification
    div_mc, peaks = parse_division_mc_and_peaks(wb, config)
    peak_sum = sum(peaks.values())
    print(f"  Division MC + peaks parsed (Σ peaks = {peak_sum:.2f} MW)")

    # 5. Combine upstream undiluted using capacity weights
    upstream_undiluted = combine_upstream_undiluted(
        undiluted_raw["ups_sub"], undiluted_raw["ups_feed"], cap
    )
    undiluted_inc: dict[str, list[float]] = {
        "upstream": upstream_undiluted,
        "dist_sub": undiluted_raw["dist_sub"],
        "primary_feeder": undiluted_raw["primary_feeder"],
        "total": undiluted_raw["total"],
    }

    # 6. Cumulative variants
    cum_diluted = accumulate_cumulative_diluted(diluted_inc)
    cum_undiluted = accumulate_cumulative_undiluted(undiluted_inc, cap)

    # 7. Table 2 derivation for verification
    table2_derived = derive_table2(
        div_mc, peaks, config.system_peak_mw, config.has_division_t5_t6
    )

    variants = {
        "incremental_diluted": diluted_inc,
        "cumulative_diluted": cum_diluted,
        "incremental_undiluted": undiluted_inc,
        "cumulative_undiluted": cum_undiluted,
    }

    # 8. Export CSVs
    print("\nExporting CSVs:")
    for vname in VARIANT_NAMES:
        prefix = f"{config.utility}_{vname}"
        export_annualized_csv(variants[vname], output_dir / f"{prefix}_annualized.csv")
        export_levelized_csv(variants[vname], output_dir / f"{prefix}_levelized.csv")

    # 9. Report
    print()
    print_report(
        config,
        variants,
        table2_derived,
        table2_wb,
        diluted_lev_wb,
        undiluted_lev_wb,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="NYSEG MCOS analysis (CRA)")
    parser.add_argument("--path-xlsx", required=True)
    parser.add_argument("--system-peak-mw", type=float, required=True)
    parser.add_argument("--path-output-dir", required=True)
    args = parser.parse_args()

    config = CRAConfig(
        utility="nyseg",
        display_name="NYSEG",
        divisions=NYSEG_DIVISIONS,
        system_peak_mw=args.system_peak_mw,
        t4_pattern="T4 Summary",
        t4b_pattern="T4B",
        has_division_t5_t6=True,
    )

    run_analysis(args.path_xlsx, config, Path(args.path_output_dir))


if __name__ == "__main__":
    main()
