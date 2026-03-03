"""Compute marginal costs from Con Edison's MCOS study workbook.

Reads ConEd's 2025 MCOS workbook (CapEx sheets for capital budgets, Schedule 11
for composite rates and escalation, Coincident Load for system peak) and computes
marginal costs by cost center in four variants:

  1. Cumulative diluted    — MCOS perspective: accumulated capital ÷ system peak
  2. Incremental diluted   — BAT perspective: new capital per year ÷ system peak
  3. Cumulative undiluted  — accumulated capital ÷ cumulative project capacity
  4. Incremental undiluted — new capital per year ÷ new capacity per year

Each variant is exported in both annualized (year-by-year) and levelized form,
for a total of 8 CSVs.

ConEd's workbook separates costs into five NERA cost centers:
  1. Transmission (138/345 kV) — NYISO-jurisdictional bulk TX, EXCLUDED
  2. Area Station & Sub-Transmission — sub-TX + dist substations, INCLUDED
  3. Primary — distribution feeders, INCLUDED
  4. Transformer — distribution transformers, INCLUDED
  5. Secondary — secondary cables, INCLUDED

Cost centers 1–2 use a cumulative 10-year capital budget (CapEx Transmission,
CapEx Substation sheets). Cost centers 3–5 use a representative annual budget
from a single sample year (CapEx Distribution sheet).

MC formula:
  Annual RR(Y) = Capital(Y) × Composite Rate × Escalation(Y)
  MC(Y)        = Annual RR(Y) / Denominator   [$/kW-yr]

  Denominator = system peak (MW) for diluted, project capacity (MW) for undiluted.
  Capital = cumulative in-service total for cumulative variants,
            year's new addition for incremental variants.

The composite rate is Schedule 11 col 13 ("Annual MC at System Peak"), which
already adjusts for area-station-to-system diversity. Dividing by the area
station coincident total (not the lower system coincident peak) is correct:
  (Capital × col13) / ASC_total  ≡  (Capital × col11) / (ASC_total × CF)
                                 ≡  Annual RR / System Peak

Classification: ConEd's own cost center structure IS the tier classification.
CapEx Transmission = bulk TX (verified against NYISO Gold Book Table VII —
both projects appear: Eastern Queens 138 kV and Brooklyn Clean Energy Hub
345 kV). No project-level reclassification is needed.

Outputs (8 CSVs):
  - coned_{cumulative,incremental}_{diluted,undiluted}_{levelized,annualized}.csv
  - Terminal report (always printed)

Usage (via Justfile):
    just analyze-coned
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import fsspec
import openpyxl
import polars as pl


# ── Study parameters ─────────────────────────────────────────────────────────

YEARS = list(range(2025, 2035))  # 2025 through 2034
N_YEARS = len(YEARS)

COST_CENTERS = ["transmission", "substation", "primary", "transformer", "secondary"]
LOCAL_CENTERS = ["substation", "primary", "transformer", "secondary"]
DIST_CENTERS = ["primary", "transformer", "secondary"]

CC_LABELS = {
    "transmission": "Bulk TX (138/345 kV)",
    "substation": "Area Station & Sub-TX",
    "primary": "Primary Distribution",
    "transformer": "Distribution Transformer",
    "secondary": "Secondary Cable",
    "distribution": "Distribution (Prim+Trans+Sec)",
    "local_total": "Sub-TX + dist (excl. bulk TX)",
}

VARIANT_NAMES = [
    "cumulative_diluted",
    "incremental_diluted",
    "cumulative_undiluted",
    "incremental_undiluted",
]


# ── ConEd workbook cell references ───────────────────────────────────────────
#
# Column/row indices are openpyxl 1-indexed (A=1, B=2, ...).
# Verified from direct inspection of coned_study_workpaper.xlsx.

# Sheet names.  Note: Carrying Charge Loaders has a TRAILING SPACE.
SH_TX = "CapEx Transmission"
SH_SUB = "CapEx Substation"
SH_DIST = "CapEx Distribution"
SH_SCHED11 = "Carrying Charge Loaders "
SH_COINC = "Coincident Load"

# ── CapEx Transmission / Substation: shared column layout ────────────────────
#
# These sheets have a left half ($/kW view) and right half (project detail).
# Both halves share the same rows; the years 2025–2034 appear twice (left for
# $/kW, right for cumulative cashflow $000s).
#
# Left half (Section 1 and Section 2 reuse these columns):
#   B(2)=MW  D(4)=Region  E(5)=Area Substation  F(6)..O(15)=years 2025–2034
#
# Right half (Section 1 only):
#   Q(17)=Area Station Ref  R(18)=Description  S(19)=Additional Capacity (MW)
#   T(20)=Estimated Cost ($000s)  U(21)=$/kW  V(22)=Pre-2025 Actual
#   W(23)..AF(32)=cumulative cashflow 2025–2034 ($000s)

CAPEX_YEAR_COL = {yr: 6 + (yr - 2025) for yr in YEARS}  # F(6)=2025 .. O(15)=2034
CAPEX_MW_COL = 2  # B: MW (left half, Section 1)
CAPEX_REGION_COL = 4  # D: Region
CAPEX_STNNAME_COL = 5  # E: Area Substation name
CAPEX_REF_COL = 17  # Q: Area Station Ref (right half)
CAPEX_PROJECT_MW_COL = 19  # S: Additional Capacity (MW, right half)

# Right-half cumulative cashflow columns ($000s)
CAPEX_CF_YEAR_COL = {yr: 23 + (yr - 2025) for yr in YEARS}  # W(23)=2025..AF(32)=2034

# Section 1 project data rows
TX_PROJECT_ROWS = range(8, 13)  # rows 8–12 (5 area-stn assignments, 2 projects)
SUB_PROJECT_ROWS = range(9, 26)  # rows 9–25 (17 area-stn projects)

# ── CapEx Distribution ───────────────────────────────────────────────────────
#
# Right-side table (cols I–O) lists individual demand-related distribution
# projects from a representative sample year.  Values are in DOLLARS (not
# $000s), unlike the other CapEx sheets.
#
# Row 151 is the total.  Project rows run from 8 to ~150.
DIST_TOTAL_ROW = 151
DIST_PRIMARY_COL = 12  # L: Primary cost ($)
DIST_TRANS_COL = 13  # M: Transformer cost ($)
DIST_SEC_COL = 14  # N: Secondary cost ($)
DIST_KW_COL = 15  # O: New capacity (kW)
DIST_NAME_COL = 9  # I: Project type / name

# ── Schedule 11: Carrying Charge Loaders ─────────────────────────────────────
#
# Composite rate at system peak level = col O (15), i.e. Schedule 11 col (13).
# GDP escalation rate = row 25, year columns C(3)..L(12).
SCHED11_RATE_COL = 15
SCHED11_ROWS = {
    "transmission": 12,
    "substation": 13,
    "primary": 14,
    "transformer": 15,
    "secondary": 16,
}
SCHED11_ESC_ROW = 25
SCHED11_ESC_YEAR_COL = {yr: 3 + (yr - 2025) for yr in YEARS}

# ── Coincident Load ──────────────────────────────────────────────────────────
#
# "Area Station Coincident Totals" row 26, year cols B(2)..L(12) for 2025–2035.
COINC_TOTAL_ROW = 26
COINC_YEAR_COL = {yr: 2 + (yr - 2025) for yr in YEARS}


# ── Data classes ─────────────────────────────────────────────────────────────


@dataclass
class ProjectData:
    """Per-project data parsed from right-half cashflow columns."""

    name: str
    capacity_mw: float
    final_capital_k: float  # cumulative capital at completion ($000s)
    in_service_year: int  # first year where cashflow = final (stabilization)


@dataclass
class CostCenterData:
    """Parsed data for one cost center."""

    name: str
    label: str
    n_projects: int
    total_capacity_mw: float
    total_cost_k: float  # total (or annual) capital $000s
    composite_rate: float
    capital_by_year: dict[int, float] = field(default_factory=dict)
    capacity_by_year: dict[int, float] = field(default_factory=dict)
    is_annual: bool = False
    projects: list[ProjectData] = field(default_factory=list)


@dataclass
class MCRow:
    """One year's MC for a cost center (used for all variants)."""

    year: int
    capital_k: float
    annual_rr_k: float
    escalation: float
    nominal_mc: float  # with escalation ($/kW-yr)
    real_mc: float  # base-year dollars ($/kW-yr)


# ── Workbook I/O ─────────────────────────────────────────────────────────────


def _cell(sheet, row: int, col: int) -> float:
    v = sheet.cell(row=row, column=col).value
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _str_cell(sheet, row: int, col: int) -> str:
    v = sheet.cell(row=row, column=col).value
    return str(v).strip() if v else ""


def _load_wb(path: str):
    if path.startswith("s3://"):
        fs = fsspec.filesystem("s3")
        with fs.open(path, "rb") as f:
            return openpyxl.load_workbook(f, data_only=True)
    return openpyxl.load_workbook(path, data_only=True)


# ── Parsing functions ────────────────────────────────────────────────────────


def parse_composite_rates(wb) -> dict[str, float]:
    sheet = wb[SH_SCHED11]
    rates = {}
    for name, row in SCHED11_ROWS.items():
        rates[name] = _cell(sheet, row, SCHED11_RATE_COL)
    return rates


def parse_escalation(wb) -> dict[int, float]:
    sheet = wb[SH_SCHED11]
    return {
        yr: _cell(sheet, SCHED11_ESC_ROW, col)
        for yr, col in SCHED11_ESC_YEAR_COL.items()
    }


def parse_coincident_peak(wb) -> dict[int, float]:
    sheet = wb[SH_COINC]
    return {
        yr: _cell(sheet, COINC_TOTAL_ROW, col) for yr, col in COINC_YEAR_COL.items()
    }


def _detect_in_service_year(cashflows: list[float]) -> int:
    """Infer in-service year from cumulative cashflow stabilization.

    Returns the first study year where the cashflow equals its final value,
    indicating CWIP has ended and the project is complete.
    Falls back to the last study year if cashflow never fully stabilizes.
    """
    final = cashflows[-1]
    if final <= 0:
        return YEARS[-1]
    for i, v in enumerate(cashflows):
        if abs(v - final) < 0.5:
            return YEARS[i]
    return YEARS[-1]


def _parse_capex_cumulative(
    wb, sheet_name: str, project_rows: range, name: str
) -> CostCenterData:
    """Parse a cumulative CapEx sheet using per-project cashflow data.

    Reads each project's right-half cumulative cashflow columns to detect
    in-service year (first year where cashflow stabilizes at its final value).
    Then builds capital_by_year and capacity_by_year using in-service-year
    scoping, matching the NiMo/CenHud methodology:

      capital(Y) = sum(p.final_capital for p where p.in_service_year <= Y)
      capacity(Y) = sum(p.capacity_mw for p where p.in_service_year <= Y)

    This excludes CWIP (construction-work-in-progress) from early years and
    ensures each project's capital and capacity enter the MC calculation
    together when the project completes.
    """
    sheet = wb[sheet_name]
    projects: list[ProjectData] = []

    for r in project_rows:
        ref = _str_cell(sheet, r, CAPEX_REF_COL)
        if not ref:
            continue
        mw = _cell(sheet, r, CAPEX_PROJECT_MW_COL)
        cashflows = [_cell(sheet, r, col) for col in CAPEX_CF_YEAR_COL.values()]
        in_svc = _detect_in_service_year(cashflows)
        final_cap = cashflows[-1]
        projects.append(ProjectData(ref, mw, final_cap, in_svc))

    capital: dict[int, float] = {}
    capacity: dict[int, float] = {}
    for yr in YEARS:
        in_scope = [p for p in projects if p.in_service_year <= yr]
        capital[yr] = sum(p.final_capital_k for p in in_scope)
        capacity[yr] = sum(p.capacity_mw for p in in_scope)

    total_mw = sum(p.capacity_mw for p in projects)
    total_cost = sum(p.final_capital_k for p in projects)

    return CostCenterData(
        name=name,
        label=CC_LABELS[name],
        n_projects=len({p.name for p in projects}),
        total_capacity_mw=total_mw,
        total_cost_k=total_cost,
        composite_rate=0.0,
        capital_by_year=capital,
        capacity_by_year=capacity,
        projects=projects,
    )


def parse_capex_tx(wb) -> CostCenterData:
    return _parse_capex_cumulative(wb, SH_TX, TX_PROJECT_ROWS, "transmission")


def parse_capex_sub(wb) -> CostCenterData:
    return _parse_capex_cumulative(wb, SH_SUB, SUB_PROJECT_ROWS, "substation")


def parse_capex_dist(wb) -> tuple[CostCenterData, CostCenterData, CostCenterData]:
    """Parse CapEx Distribution for the three distribution cost centers.

    CapEx Distribution values are in DOLLARS (not $000s).  We convert to $000s
    for consistency with the other CapEx sheets.

    Annual cost centers are already incremental: each year repeats the same
    sample.  capacity_by_year is flat (same MW every year).
    """
    sheet = wb[SH_DIST]

    primary_dollars = _cell(sheet, DIST_TOTAL_ROW, DIST_PRIMARY_COL)
    trans_dollars = _cell(sheet, DIST_TOTAL_ROW, DIST_TRANS_COL)
    sec_dollars = _cell(sheet, DIST_TOTAL_ROW, DIST_SEC_COL)
    total_kw = _cell(sheet, DIST_TOTAL_ROW, DIST_KW_COL)
    total_mw = total_kw / 1000.0

    n_projects = 0
    for r in range(8, DIST_TOTAL_ROW):
        if _str_cell(sheet, r, DIST_NAME_COL):
            n_projects += 1

    def _make(name: str, dollars: float) -> CostCenterData:
        cap_k = dollars / 1000.0
        return CostCenterData(
            name=name,
            label=CC_LABELS[name],
            n_projects=n_projects,
            total_capacity_mw=total_mw,
            total_cost_k=cap_k,
            composite_rate=0.0,
            capital_by_year={yr: cap_k * (i + 1) for i, yr in enumerate(YEARS)},
            capacity_by_year={yr: total_mw * (i + 1) for i, yr in enumerate(YEARS)},
            is_annual=True,
        )

    return (
        _make("primary", primary_dollars),
        _make("transformer", trans_dollars),
        _make("secondary", sec_dollars),
    )


# ── MC computation ───────────────────────────────────────────────────────────


def incremental_from_cumulative(by_year: dict[int, float]) -> dict[int, float]:
    """Convert cumulative year-by-year values to year-over-year deltas."""
    result: dict[int, float] = {}
    prev = 0.0
    for yr in YEARS:
        val = by_year.get(yr, 0.0)
        result[yr] = val - prev
        prev = val
    return result


def compute_mc(
    capital_by_year: dict[int, float],
    composite_rate: float,
    escalation: dict[int, float],
    denominator_by_year: dict[int, float],
) -> list[MCRow]:
    """Year-by-year MC for one cost center.

    capital is in $000s, denominator in MW ⇒ $000s/MW = $/kW.
    When denominator is 0 (no capacity in that year), MC is 0.
    """
    rows: list[MCRow] = []
    for yr in YEARS:
        cap = capital_by_year.get(yr, 0.0)
        esc = escalation.get(yr, 1.0)
        denom = denominator_by_year.get(yr, 0.0)
        rr_nom = cap * composite_rate * esc
        rr_real = cap * composite_rate
        nom_mc = rr_nom / denom if denom > 0 else 0.0
        real_mc = rr_real / denom if denom > 0 else 0.0
        rows.append(MCRow(yr, cap, rr_nom, esc, nom_mc, real_mc))
    return rows


def levelized(rows: list[MCRow]) -> float:
    """Average of real (base-year) MC across all study years."""
    return sum(r.real_mc for r in rows) / len(rows) if rows else 0.0


# ── CSV export ───────────────────────────────────────────────────────────────


def _sum_mc(
    mc_data: dict[str, list[MCRow]], centers: list[str], yi: int, *, nominal: bool
) -> float:
    return sum(
        (mc_data[c][yi].nominal_mc if nominal else mc_data[c][yi].real_mc)
        for c in centers
    )


def export_levelized_csv(
    ccs: dict[str, CostCenterData],
    mc_data: dict[str, list[MCRow]],
    path: Path,
) -> None:
    rows = []
    for name in COST_CENTERS:
        cc = ccs[name]
        lev = levelized(mc_data[name])
        full_real = mc_data[name][-1].real_mc
        full_nom = mc_data[name][-1].nominal_mc
        cost_m = (
            cc.total_cost_k / 1e3
            if cc.is_annual
            else cc.capital_by_year[YEARS[-1]] / 1e3
        )
        rows.append(
            {
                "cost_center": name,
                "label": cc.label,
                "n_projects": cc.n_projects,
                "capacity_mw": round(cc.total_capacity_mw, 1),
                "total_cost_m": round(cost_m, 1),
                "composite_rate": round(cc.composite_rate, 5),
                "levelized_mc_kw_yr": round(lev, 2),
                "final_year_real_mc_kw_yr": round(full_real, 2),
                "final_year_nominal_mc_kw_yr": round(full_nom, 2),
            }
        )

    dist_lev = sum(levelized(mc_data[c]) for c in DIST_CENTERS)
    dist_full_real = sum(mc_data[c][-1].real_mc for c in DIST_CENTERS)
    rows.append(
        {
            "cost_center": "distribution",
            "label": CC_LABELS["distribution"],
            "n_projects": ccs["primary"].n_projects,
            "capacity_mw": round(ccs["primary"].total_capacity_mw, 1),
            "total_cost_m": round(
                sum(ccs[c].total_cost_k for c in DIST_CENTERS) / 1e3, 1
            ),
            "composite_rate": 0.0,
            "levelized_mc_kw_yr": round(dist_lev, 2),
            "final_year_real_mc_kw_yr": round(dist_full_real, 2),
            "final_year_nominal_mc_kw_yr": round(
                sum(mc_data[c][-1].nominal_mc for c in DIST_CENTERS), 2
            ),
        }
    )

    local_lev = sum(levelized(mc_data[c]) for c in LOCAL_CENTERS)
    local_full_real = sum(mc_data[c][-1].real_mc for c in LOCAL_CENTERS)
    rows.append(
        {
            "cost_center": "local_total",
            "label": CC_LABELS["local_total"],
            "n_projects": 0,
            "capacity_mw": 0.0,
            "total_cost_m": 0.0,
            "composite_rate": 0.0,
            "levelized_mc_kw_yr": round(local_lev, 2),
            "final_year_real_mc_kw_yr": round(local_full_real, 2),
            "final_year_nominal_mc_kw_yr": round(
                sum(mc_data[c][-1].nominal_mc for c in LOCAL_CENTERS), 2
            ),
        }
    )

    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


def export_annualized_csv(mc_data: dict[str, list[MCRow]], path: Path) -> None:
    rows = []
    for yi, yr in enumerate(YEARS):
        row: dict[str, object] = {"year": yr}
        for name in COST_CENTERS:
            row[f"{name}_nominal"] = round(mc_data[name][yi].nominal_mc, 2)
            row[f"{name}_real"] = round(mc_data[name][yi].real_mc, 2)
        row["distribution_nominal"] = round(
            _sum_mc(mc_data, DIST_CENTERS, yi, nominal=True), 2
        )
        row["distribution_real"] = round(
            _sum_mc(mc_data, DIST_CENTERS, yi, nominal=False), 2
        )
        row["local_total_nominal"] = round(
            _sum_mc(mc_data, LOCAL_CENTERS, yi, nominal=True), 2
        )
        row["local_total_real"] = round(
            _sum_mc(mc_data, LOCAL_CENTERS, yi, nominal=False), 2
        )
        rows.append(row)
    pl.DataFrame(rows).write_csv(path)
    print(f"  Wrote {path}")


# ── Terminal report ──────────────────────────────────────────────────────────


def print_report(
    ccs: dict[str, CostCenterData],
    variants: dict[str, dict[str, list[MCRow]]],
    system_peak_mw: float,
) -> None:
    W = 80
    print("=" * W)
    print("Con Edison MCOS Analysis — Cumulative vs. Incremental")
    print("=" * W)

    print(f"\n── System {'─' * (W - 11)}")
    print(f"  System peak (2025 ASC total):  {system_peak_mw:,.1f} MW")

    # Project in-service year summary for cumulative cost centers
    for name in ["transmission", "substation"]:
        cc = ccs[name]
        if cc.projects:
            print(
                f"\n── {cc.label} — in-service years {'─' * max(1, W - 35 - len(cc.label))}"
            )
            by_year: dict[int, list[ProjectData]] = {}
            for p in cc.projects:
                by_year.setdefault(p.in_service_year, []).append(p)
            for yr in sorted(by_year):
                projs = by_year[yr]
                names = ", ".join(p.name for p in projs)
                mw = sum(p.capacity_mw for p in projs)
                cap = sum(p.final_capital_k for p in projs)
                print(
                    f"  {yr}: {len(projs)} project(s), {mw:,.0f} MW, ${cap / 1e3:,.0f}M — {names}"
                )

    cum_dil = variants["cumulative_diluted"]
    inc_dil = variants["incremental_diluted"]
    cum_undil = variants["cumulative_undiluted"]
    inc_undil = variants["incremental_undiluted"]

    print(f"\n── Levelized MC ($/kW-yr, real) {'─' * (W - 32)}")
    print(
        f"  {'Cost center':<28} {'Cum.Dil':>8} {'Inc.Dil':>8}"
        f" {'Cum.Und':>8} {'Inc.Und':>8}"
    )
    print(f"  {'─' * (W - 4)}")
    for name in COST_CENTERS:
        cc = ccs[name]
        vals = [
            levelized(cum_dil[name]),
            levelized(inc_dil[name]),
            levelized(cum_undil[name]),
            levelized(inc_undil[name]),
        ]
        parts = "  ".join(f"${v:>6.2f}" for v in vals)
        print(f"  {cc.label:<28} {parts}")
    loc_vals = [
        sum(levelized(v[c]) for c in LOCAL_CENTERS)
        for v in [cum_dil, inc_dil, cum_undil, inc_undil]
    ]
    parts = "  ".join(f"${v:>6.2f}" for v in loc_vals)
    print(f"  {'─' * (W - 4)}")
    print(f"  {'Sub-TX + dist total':<28} {parts}")

    print(f"\n── Year-by-year local_total diluted ($/kW-yr, nominal) {'─' * (W - 55)}")
    print(f"  {'Year':>4}  {'Cumulative':>12}  {'Incremental':>12}  {'Match?':>8}")
    print(f"  {'─' * 42}")
    for yi, yr in enumerate(YEARS):
        cum_v = _sum_mc(cum_dil, LOCAL_CENTERS, yi, nominal=True)
        inc_v = _sum_mc(inc_dil, LOCAL_CENTERS, yi, nominal=True)
        match = "  ✓" if yr == YEARS[0] and abs(cum_v - inc_v) < 0.01 else ""
        print(f"  {yr:>4}  ${cum_v:>10.2f}  ${inc_v:>10.2f}  {match}")

    print()


# ── CLI ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compute MC variants from ConEd MCOS workbook"
    )
    parser.add_argument(
        "--path-xlsx",
        type=str,
        required=True,
        help="Path or S3 URL to ConEd MCOS workbook",
    )
    parser.add_argument(
        "--system-peak-mw",
        type=float,
        required=True,
        help="Area station coincident total (MW)",
    )
    parser.add_argument(
        "--path-output-dir", type=Path, required=True, help="Directory for output CSVs"
    )
    args = parser.parse_args()

    print(f"Loading workbook: {args.path_xlsx}")
    wb = _load_wb(args.path_xlsx)

    composite = parse_composite_rates(wb)
    escalation = parse_escalation(wb)
    coinc = parse_coincident_peak(wb)
    print(
        f"  Composite rates: {', '.join(f'{k}={v:.5f}' for k, v in composite.items())}"
    )
    print(f"  Escalation: {escalation[2025]:.4f} → {escalation[2034]:.4f}")
    print(
        f"  Coincident load 2025: {coinc[2025]:,.1f} MW (CLI: {args.system_peak_mw:,.1f} MW)"
    )

    tx = parse_capex_tx(wb)
    sub = parse_capex_sub(wb)
    primary, transformer, secondary = parse_capex_dist(wb)

    ccs: dict[str, CostCenterData] = {}
    for cc in [tx, sub, primary, transformer, secondary]:
        cc.composite_rate = composite[cc.name]
        ccs[cc.name] = cc

    tx_cap_m = tx.capital_by_year[YEARS[-1]] / 1e3
    sub_cap_m = sub.capital_by_year[YEARS[-1]] / 1e3
    dist_k = primary.total_cost_k + transformer.total_cost_k + secondary.total_cost_k
    print(
        f"  TX: {tx.n_projects} project(s), {tx.total_capacity_mw:,.0f} MW, ${tx_cap_m:,.0f}M capital"
    )
    print(
        f"  Sub: {sub.n_projects} project(s), {sub.total_capacity_mw:,.0f} MW, ${sub_cap_m:,.0f}M capital"
    )
    print(
        f"  Dist: {primary.n_projects} projects (sample), ${dist_k / 1e3:,.1f}M annual"
    )

    # ── Build all 4 variants ──────────────────────────────────────────────────

    peak_denom = {yr: args.system_peak_mw for yr in YEARS}

    # Incremental capital and capacity (year-over-year deltas of cumulative)
    inc_capital: dict[str, dict[int, float]] = {}
    inc_capacity: dict[str, dict[int, float]] = {}
    for name, cc in ccs.items():
        inc_capital[name] = incremental_from_cumulative(cc.capital_by_year)
        inc_capacity[name] = incremental_from_cumulative(cc.capacity_by_year)

    variants: dict[str, dict[str, list[MCRow]]] = {}

    variants["cumulative_diluted"] = {
        name: compute_mc(cc.capital_by_year, cc.composite_rate, escalation, peak_denom)
        for name, cc in ccs.items()
    }

    variants["incremental_diluted"] = {
        name: compute_mc(
            inc_capital[name], ccs[name].composite_rate, escalation, peak_denom
        )
        for name in COST_CENTERS
    }

    variants["cumulative_undiluted"] = {
        name: compute_mc(
            cc.capital_by_year, cc.composite_rate, escalation, cc.capacity_by_year
        )
        for name, cc in ccs.items()
    }

    variants["incremental_undiluted"] = {
        name: compute_mc(
            inc_capital[name],
            ccs[name].composite_rate,
            escalation,
            inc_capacity[name],
        )
        for name in COST_CENTERS
    }

    # ── Export CSVs ───────────────────────────────────────────────────────────

    out = args.path_output_dir
    out.mkdir(parents=True, exist_ok=True)
    print("\nExporting CSVs:")
    for variant_name in VARIANT_NAMES:
        mc_data = variants[variant_name]
        prefix = f"coned_{variant_name}"
        export_levelized_csv(ccs, mc_data, out / f"{prefix}_levelized.csv")
        export_annualized_csv(mc_data, out / f"{prefix}_annualized.csv")

    print()
    print_report(ccs, variants, args.system_peak_mw)

    wb.close()


if __name__ == "__main__":
    main()
