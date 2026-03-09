#!/usr/bin/env python3
"""Fetch ISO-NE CELT Excel files and extract Table 4.1 (Summary of CSOs) to parquet.

Downloads CELT reports from ISO-NE, parses the CSO sheet, and writes Hive-partitioned
parquet (celt_year={YYYY}/data.parquet).

Usage:
    uv run python data/isone/capacity/cso/fetch_isone_celt_cso.py \
        --all --path-local-parquet parquet --path-local-xlsx xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import openpyxl.worksheet.worksheet
import polars as pl
import urllib.request

CELT_URLS: dict[int, str] = {
    2025: "https://www.iso-ne.com/static-assets/documents/100023/2025_celt.xlsx",
    2023: "https://www.iso-ne.com/static-assets/documents/2023/05/2023_celt_report.xlsx",
    2022: "https://www.iso-ne.com/static-assets/documents/2022/04/2022_celt_report.xlsx",
    2021: "https://www.iso-ne.com/static-assets/documents/2021/04/2021_celt_report.xlsx",
    2020: "https://www.iso-ne.com/static-assets/documents/2020/04/2020_celt_report.xlsx",
}

SHEET_NAMES = ["4.1 Summary of CSOs", "4.1 Sum of CSOs"]

LOAD_ZONES = {"CT", "ME", "NEMA", "NH", "RI", "SEMA", "VT", "WCMA"}

# CCP date format changed: "2019-2020" (2020-2022 CELTs) → "2024/2025" (2023+ CELTs)
# Also uses en-dash: "2022–2023"
CCP_DATE_RE = re.compile(r"(\d{4})[/–-](\d{4})")
FCA_RE = re.compile(r"FCA\s+(\d+)")

# Mapping from (resource_type, resource_subtype) in Excel → our canonical resource_type.
# We extract rows at the subtotal and detail level we need.
RESOURCE_TYPE_MAP: dict[str, str] = {
    "TOTAL ACTIVE": "Active DCR",
    "TOTAL PASSIVE": "Passive DCR",
    "Intermittent": "Gen Intermittent",
    "Non Intermittent": "Gen Non-Intermittent",
    "Non-Intermittent": "Gen Non-Intermittent",
    "DR Total": "DCR Total",
    "DCR Total": "DCR Total",
    "GEN Total": "Gen Total",
}


def download_celt(year: int, xlsx_dir: Path) -> Path:
    """Download a CELT Excel file, returning local path. Skip if already cached."""
    url = CELT_URLS[year]
    filename = url.rsplit("/", 1)[-1]
    local_path = xlsx_dir / filename
    if local_path.exists():
        print(f"  [cached] {local_path.name}")
        return local_path
    print(f"  Downloading {url} ...")
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(url, local_path)
    print(f"  Saved to {local_path.name}")
    return local_path


def find_cso_sheet(
    wb: openpyxl.Workbook,
) -> openpyxl.worksheet.worksheet.Worksheet | None:
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if "4.1" in name:
            return wb[name]
    return None


def parse_ccp_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
) -> list[tuple[int, str, int | None]]:
    """Parse CCP headers from the merged header row.

    Returns list of (col_index_for_summer, ccp_string, fca_number_or_none).
    Summer CSO is at col_index, Winter CSO is at col_index + 1.
    """
    ccps: list[tuple[int, str, int | None]] = []
    for col_idx in range(5, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        val_str = str(val).strip()
        ccp_match = CCP_DATE_RE.search(val_str)
        if not ccp_match:
            continue
        ccp = f"{ccp_match.group(1)}/{ccp_match.group(2)}"
        fca_match = FCA_RE.search(val_str)
        fca_number = int(fca_match.group(1)) if fca_match else None
        ccps.append((col_idx, ccp, fca_number))
    return ccps


def derive_fca_number(ccp: str) -> int:
    """Derive FCA number from CCP dates. FCA 1 was for 2010/2011."""
    match = CCP_DATE_RE.match(ccp)
    if not match:
        raise ValueError(f"Cannot parse CCP: {ccp}")
    start_year = int(match.group(1))
    return start_year - 2010 + 1


def parse_cso_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    celt_year: int,
) -> pl.DataFrame:
    """Parse a single CSO sheet into a tidy DataFrame."""
    # Find the header row: scan for "Capacity Commitment Period" in column 5
    header_row = None
    for row_idx in range(1, 20):
        val = ws.cell(row=row_idx, column=5).value
        if val and "Capacity Commitment Period" in str(val):
            header_row = row_idx
            break
    if header_row is None:
        raise ValueError(f"CELT {celt_year}: could not find CCP header row")

    ccp_header_row = header_row + 1
    ccps = parse_ccp_headers(ws, ccp_header_row)
    if not ccps:
        raise ValueError(f"CELT {celt_year}: no CCPs found in row {ccp_header_row}")

    print(f"  Found {len(ccps)} CCPs: {[c[1] for c in ccps]}")

    # Assign FCA numbers: use explicit FCA from header, fall back to derivation
    ccp_info: list[tuple[int, str, int]] = []
    for col_idx, ccp, fca in ccps:
        if fca is None:
            fca = derive_fca_number(ccp)
        ccp_info.append((col_idx, ccp, fca))

    data_start_row = ccp_header_row + 2
    rows: list[dict] = []
    current_zone: str | None = None

    for row_idx in range(data_start_row, ws.max_row + 1):
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        col_d = ws.cell(row=row_idx, column=4).value

        col_b_str = str(col_b).strip() if col_b is not None else ""
        col_c_str = str(col_c).strip() if col_c is not None else ""
        col_d_str = str(col_d).strip() if col_d is not None else ""

        # Detect zone from column B
        if col_b_str:
            zone_candidate = col_b_str.upper().replace(" ", "")
            if zone_candidate in LOAD_ZONES:
                current_zone = zone_candidate
                continue
            # "{zone} Total" row
            total_match = re.match(r"(\w+)\s+Total", col_b_str, re.IGNORECASE)
            if total_match:
                zone_name = total_match.group(1).strip().upper().replace(" ", "")
                if zone_name in LOAD_ZONES:
                    resource_type = "Total"
                    for col_idx, ccp, fca in ccp_info:
                        summer = ws.cell(row=row_idx, column=col_idx).value
                        winter = ws.cell(row=row_idx, column=col_idx + 1).value
                        rows.append(
                            {
                                "celt_year": celt_year,
                                "fca_number": fca,
                                "ccp": ccp,
                                "state": zone_name,
                                "resource_type": resource_type,
                                "summer_cso_mw": _to_float(summer),
                                "winter_cso_mw": _to_float(winter),
                            }
                        )
                    continue
            # ISO-level rows — skip
            if "ISO" in col_b_str or "Grand" in col_b_str or "Import" in col_b_str:
                current_zone = None
                continue

        if current_zone is None:
            continue

        # Determine the canonical resource type for this row
        canonical = None

        # Column C has the resource type label for subtotal rows
        if col_c_str in RESOURCE_TYPE_MAP:
            canonical = RESOURCE_TYPE_MAP[col_c_str]
        elif col_d_str in RESOURCE_TYPE_MAP:
            canonical = RESOURCE_TYPE_MAP[col_d_str]

        if canonical is None:
            continue

        for col_idx, ccp, fca in ccp_info:
            summer = ws.cell(row=row_idx, column=col_idx).value
            winter = ws.cell(row=row_idx, column=col_idx + 1).value
            rows.append(
                {
                    "celt_year": celt_year,
                    "fca_number": fca,
                    "ccp": ccp,
                    "state": current_zone,
                    "resource_type": canonical,
                    "summer_cso_mw": _to_float(summer),
                    "winter_cso_mw": _to_float(winter),
                }
            )

    if not rows:
        raise ValueError(f"CELT {celt_year}: parsed zero data rows")

    df = pl.DataFrame(rows).cast(
        {
            "celt_year": pl.Int16,
            "fca_number": pl.Int16,
            "ccp": pl.String,
            "state": pl.String,
            "resource_type": pl.String,
            "summer_cso_mw": pl.Float64,
            "winter_cso_mw": pl.Float64,
        }
    )
    return df


def _to_float(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def process_celt(year: int, xlsx_dir: Path) -> pl.DataFrame:
    """Download, parse, and return DataFrame for a single CELT year."""
    print(f"\nProcessing CELT {year}...")
    xlsx_path = download_celt(year, xlsx_dir)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = find_cso_sheet(wb)
    if ws is None:
        wb.close()
        raise ValueError(f"CELT {year}: no CSO sheet found. Sheets: {wb.sheetnames}")
    print(f"  Sheet: {ws.title}")
    df = parse_cso_sheet(ws, year)
    wb.close()
    zones = sorted(df["state"].unique().to_list())
    print(f"  Parsed {df.height} rows, zones: {zones}")
    return df


def main() -> int:
    parser = argparse.ArgumentParser(description="Fetch ISO-NE CELT CSO data.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--celt-years",
        type=str,
        help="Comma-separated CELT years (e.g. 2020,2021,2025).",
    )
    group.add_argument("--all", action="store_true", help="Fetch all known CELT years.")
    parser.add_argument(
        "--path-local-parquet",
        type=Path,
        required=True,
        help="Output directory for Hive-partitioned parquet.",
    )
    parser.add_argument(
        "--path-local-xlsx",
        type=Path,
        required=True,
        help="Cache directory for downloaded Excel files.",
    )
    args = parser.parse_args()

    if "{{" in str(args.path_local_parquet) or "{{" in str(args.path_local_xlsx):
        print("ERROR: path contains uninterpolated Just variable", file=sys.stderr)
        return 1

    years = (
        sorted(CELT_URLS.keys())
        if args.all
        else [int(y) for y in args.celt_years.split(",")]
    )
    unknown = [y for y in years if y not in CELT_URLS]
    if unknown:
        print(f"ERROR: unknown CELT years: {unknown}", file=sys.stderr)
        return 1

    all_dfs: list[pl.DataFrame] = []
    for year in years:
        df = process_celt(year, args.path_local_xlsx)
        all_dfs.append(df)

    combined = pl.concat(all_dfs)
    print(f"\nCombined: {combined.height} total rows")

    # Write Hive-partitioned parquet
    out_dir = args.path_local_parquet
    out_dir.mkdir(parents=True, exist_ok=True)
    for year in years:
        year_df = combined.filter(pl.col("celt_year") == year).drop("celt_year")
        year_dir = out_dir / f"celt_year={year}"
        year_dir.mkdir(parents=True, exist_ok=True)
        year_df.write_parquet(year_dir / "data.parquet")
        print(f"  Wrote {year_dir / 'data.parquet'} ({year_df.height} rows)")

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
