"""Compute the year each NY utility exhausts its distribution winter headroom.

Uses NYISO Gold Book Table I-4b (Winter Non-Coincident Peak Demand by Zone)
and utility-to-zone mappings from generate_zone_mapping_csv.py to project when
each utility's winter peak growth will consume available distribution headroom.

Supports two NYISO Gold Book scenarios:
  baseline — 2024-Gold-Book-Baseline-Forecast-Tables (2).xlsx
  lower    — 2024-Gold-Book-Lower-Demand-Scenario-Tables.xlsx
             (EE, non-solar DG, storage, and large loads are identical to
             baseline per the Notes sheet; only EV and building electrification
             differ)

Produces per-scenario outputs:
  headroom_exhaustion_year_by_utility_<scenario>.csv
  demand_components_by_utility_<scenario>.png

Source for headroom data: Table 2 from Switchbox distribution headroom analysis.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl

ZONES = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

UTILITY_ZONES: dict[str, list[str]] = {
    "National Grid": ["A", "B", "C", "D", "E", "F"],
    "Central Hudson Gas and Electric": ["G"],
    "NYS Electric & Gas and RGE": ["A", "B", "C", "D", "E", "F"],
    "Con Edison": ["G", "H", "I", "J"],
    "Orange and Rockland Utilities": ["G"],
}

# Distribution winter peak (MW) and total estimated winter headroom (MW)
# from Table 2 of the Switchbox distribution headroom analysis.
HEADROOM_TABLE: dict[str, dict[str, int | float]] = {
    "National Grid": {"dist_peak_mw": 4276, "headroom_mw": 4477, "headroom_pct": 1.05},
    "Central Hudson Gas and Electric": {
        "dist_peak_mw": 796,
        "headroom_mw": 279,
        "headroom_pct": 0.35,
    },
    "NYS Electric & Gas and RGE": {
        "dist_peak_mw": 3786,
        "headroom_mw": 3235,
        "headroom_pct": 0.85,
    },
    "Con Edison": {"dist_peak_mw": 4691, "headroom_mw": 1346, "headroom_pct": 0.29},
    "Orange and Rockland Utilities": {
        "dist_peak_mw": 1123,
        "headroom_mw": 1071,
        "headroom_pct": 0.95,
    },
}

ZoneData = dict[str, dict[str, int]]


@dataclass
class ScenarioData:
    label: str
    # I-4b / I-4b-L  — non-coincident winter peak (headroom exhaustion)
    winter_peak: ZoneData
    # I-4b from the baseline file — always has the 2023-24 historical row,
    # used as the NCP growth anchor (lower scenario I-4b-L starts at 2024-25).
    winter_peak_baseline: ZoneData
    # I-3b from the baseline file — historical coincident winter peak used for
    # normalisation; always loaded from baseline because 2023-24 is historical
    # and absent from the lower-demand scenario file.
    coincident_peak_baseline: ZoneData
    # I-13c / I-13c-L
    elec: ZoneData
    # I-11d / I-11d-L
    ev: ZoneData
    # unchanged between scenarios
    ee: ZoneData
    nondg: ZoneData
    storage: ZoneData
    large_loads: ZoneData


# ---------------------------------------------------------------------------
# Low-level parsers
# ---------------------------------------------------------------------------


def _parse_zone_sheet_winter_str(
    xlsx_path: Path,
    sheet: str,
    min_row: int = 7,
) -> ZoneData:
    """Parse a Gold Book zone-by-year table where the year column contains
    "YYYY-YY" winter season strings (e.g. "2024-25").

    Returns {year_str: {zone_letter: MW}}.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet]

    zone_data: ZoneData = {}
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        year = row[2]
        if year is None or not isinstance(year, str) or "-" not in year:
            continue
        vals = row[3:14]
        if any(v is None for v in vals):
            continue
        zone_data[year] = {z: int(v) for z, v in zip(ZONES, vals)}

    wb.close()
    return zone_data


def _parse_zone_sheet_calendar_int(
    xlsx_path: Path,
    sheet: str,
    min_row: int = 7,
) -> ZoneData:
    """Parse a Gold Book zone-by-year table where the year column is an integer
    calendar year. Returns keys normalised to "YYYY-YY" winter season strings
    matching the convention used by the other tables (year N -> "N-NN").
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet]

    zone_data: ZoneData = {}
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        year = row[2]
        if not isinstance(year, int):
            continue
        vals = row[3:14]
        if any(v is None for v in vals):
            continue
        # Map calendar year N to winter season "N-NN" (e.g. 2024 -> "2024-25")
        key = f"{year}-{(year + 1) % 100:02d}"
        zone_data[key] = {z: int(v) for z, v in zip(ZONES, vals)}

    wb.close()
    return zone_data


# ---------------------------------------------------------------------------
# Named parsers (baseline sheet names)
# ---------------------------------------------------------------------------


def parse_gold_book_winter_peak(xlsx_path: Path, suffix: str = "") -> ZoneData:
    """Parse Table I-4b[-L] (Winter Non-Coincident Peak Demand by Zone)."""
    return _parse_zone_sheet_winter_str(xlsx_path, f"I-4b{suffix}")


def parse_gold_book_coincident_winter_peak(
    xlsx_path: Path, suffix: str = ""
) -> ZoneData:
    """Parse Table I-3b[-L] (Baseline Winter Coincident Peak Demand by Zone)."""
    return _parse_zone_sheet_winter_str(xlsx_path, f"I-3b{suffix}")


def parse_gold_book_elec_load(xlsx_path: Path, suffix: str = "") -> ZoneData:
    """Parse Table I-13c[-L] (Building Electrification Winter Coincident Peak
    Demand by Zone) — cumulative future load additions in MW."""
    return _parse_zone_sheet_winter_str(xlsx_path, f"I-13c{suffix}")


def parse_gold_book_ev_load(xlsx_path: Path, suffix: str = "") -> ZoneData:
    """Parse Table I-11d[-L] (EV Winter Coincident Peak Demand by Zone)."""
    return _parse_zone_sheet_winter_str(xlsx_path, f"I-11d{suffix}")


def parse_gold_book_ee_reductions(xlsx_path: Path) -> ZoneData:
    """Parse Table I-8c (EE & Codes/Standards Winter Coincident Peak Reductions
    by Zone, relative to 2023-24) — positive values = reductions.
    Identical between baseline and lower scenario."""
    return _parse_zone_sheet_winter_str(xlsx_path, "I-8c")


def parse_gold_book_storage_reductions(xlsx_path: Path) -> ZoneData:
    """Parse Table I-12c (Storage Winter Coincident Peak Reductions by Zone).
    Identical between baseline and lower scenario."""
    return _parse_zone_sheet_calendar_int(xlsx_path, "I-12c")


def parse_gold_book_nondg_reductions(xlsx_path: Path) -> ZoneData:
    """Parse Table I-10c (Non-Solar DG Winter Coincident Peak Reductions by Zone).
    Identical between baseline and lower scenario."""
    return _parse_zone_sheet_calendar_int(xlsx_path, "I-10c")


def parse_gold_book_large_loads(xlsx_path: Path) -> ZoneData:
    """Parse the 'Winter Peak Demand by Zone' section of Table I-14.
    The table only runs through 2035-36; per the Gold Book note, forecasts for
    2036 onward equal the final year, so we forward-fill from the last row.
    Identical between baseline and lower scenario."""
    data = _parse_zone_sheet_winter_str(xlsx_path, "I-14", min_row=41)
    if not data:
        return data
    last_year = max(data.keys())
    last_values = data[last_year]
    # Forward-fill through 2054-55 to match the horizon of other tables
    start = int(last_year.split("-")[0]) + 1
    for yr in range(start, 2055):
        key = f"{yr}-{(yr + 1) % 100:02d}"
        data[key] = last_values
    return data


# ---------------------------------------------------------------------------
# Scenario loading
# ---------------------------------------------------------------------------


def load_scenario(
    path_baseline: Path,
    path_lower: Path | None = None,
) -> ScenarioData:
    """Load all zone-level data for one scenario.

    If path_lower is provided, EV and electrification come from the lower
    demand file (sheets I-11d-L and I-13c-L); all other components and the
    non-coincident winter peak (I-4b-L) also come from that file.
    The unchanged components (EE, non-solar DG, storage, large loads) always
    come from the baseline file.
    """
    if path_lower is not None:
        label = "lower"
        src = path_lower
        suffix = "-L"
    else:
        label = "baseline"
        src = path_baseline
        suffix = ""

    baseline_winter_peak = parse_gold_book_winter_peak(path_baseline)
    return ScenarioData(
        label=label,
        winter_peak=parse_gold_book_winter_peak(src, suffix),
        # always from baseline: historical 2023-24 row absent from lower file
        winter_peak_baseline=baseline_winter_peak,
        coincident_peak_baseline=parse_gold_book_coincident_winter_peak(path_baseline),
        elec=parse_gold_book_elec_load(src, suffix),
        ev=parse_gold_book_ev_load(src, suffix),
        # unchanged components always from baseline
        ee=parse_gold_book_ee_reductions(path_baseline),
        nondg=parse_gold_book_nondg_reductions(path_baseline),
        storage=parse_gold_book_storage_reductions(path_baseline),
        large_loads=parse_gold_book_large_loads(path_baseline),
    )


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------


def compute_headroom_exhaustion(
    zone_data: ZoneData,
    baseline_year: str = "2024-25",
) -> list[dict[str, str | int]]:
    """For each utility, find the first forecast year where zone-sum peak
    exceeds baseline * (1 + headroom_pct)."""
    results: list[dict[str, str | int]] = []

    for utility, zones in UTILITY_ZONES.items():
        info = HEADROOM_TABLE[utility]
        gb_baseline = sum(zone_data[baseline_year][z] for z in zones)
        threshold = gb_baseline * (1 + info["headroom_pct"])

        hit_year = None
        hit_peak = None
        last_year = None
        last_peak = None

        for year_str in sorted(zone_data.keys()):
            start_yr = int(year_str.split("-")[0])
            if start_yr < 2024:
                continue
            peak = sum(zone_data[year_str][z] for z in zones)
            if peak >= threshold and hit_year is None:
                hit_year = year_str
                hit_peak = peak
            last_year = year_str
            last_peak = peak

        results.append(
            {
                "utility": utility,
                "zones": ", ".join(zones),
                "dist_winter_peak_2024_mw": info["dist_peak_mw"],
                "headroom_mw": info["headroom_mw"],
                "headroom_pct_of_winter_peak": f"{info['headroom_pct'] * 100:.0f}%",
                "gb_2024_zone_sum_mw": gb_baseline,
                "gb_threshold_mw": round(threshold),
                "year_headroom_exhausted": hit_year
                if hit_year
                else f"Beyond {last_year}",
                "gb_peak_at_exhaustion_mw": hit_peak if hit_peak else last_peak,
            }
        )

    return results


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------


def _zone_sum_series(
    data: ZoneData,
    years: list[str],
    zones: list[str],
) -> list[float]:
    """Return zone-summed values for each year, defaulting to 0 if absent."""
    return [
        float(sum(data[yr][z] for z in zones)) if yr in data else 0.0 for yr in years
    ]


def plot_demand_components_by_utility(
    scenario: ScenarioData,
    headroom_results: list[dict[str, str | int]],
    path_output: Path,
    norm_year: str = "2023-24",
) -> None:
    """Stacked area chart of all I-1d winter coincident peak demand components
    by utility, expressed as % of the 2023-24 coincident winter peak.

    Additions (EV, building electrification, large loads) stack above zero.
    Reductions (EE, non-solar DG, storage) stack below zero.
    """
    all_years = sorted({yr for yr in scenario.elec if int(yr.split("-")[0]) >= 2024})
    x = [int(y.split("-")[0]) for y in all_years]

    exhaustion_years: dict[str, int | None] = {
        r["utility"]: (
            int(str(r["year_headroom_exhausted"]).split("-")[0])
            if not str(r["year_headroom_exhausted"]).startswith("Beyond")
            else None
        )
        for r in headroom_results
    }

    # (label, data, sign=+1 addition / -1 reduction, color)
    components: list[tuple[str, ZoneData, int, str]] = [
        ("Building electrification", scenario.elec, +1, "#e6550d"),
        ("EV load", scenario.ev, +1, "#fdae6b"),
        ("Large loads", scenario.large_loads, +1, "#756bb1"),
        ("EE & codes/standards", scenario.ee, -1, "#2ca02c"),
        ("Non-solar DG (BTM)", scenario.nondg, -1, "#74c476"),
        ("BTM storage", scenario.storage, -1, "#9edae5"),
    ]

    n = len(UTILITY_ZONES)
    fig, axes = plt.subplots(n, 1, figsize=(11, 3.6 * n), sharex=True)
    scenario_title = scenario.label.capitalize()
    fig.suptitle(
        f"Winter Coincident Peak Demand Components by Utility — {scenario_title} Scenario\n"
        f"as % of {norm_year} coincident winter peak (I-3b) — NYISO Gold Book I-1d",
        fontsize=12,
        fontweight="bold",
        y=1.01,
    )

    for ax, (utility, zones) in zip(axes, UTILITY_ZONES.items()):
        base_mw = sum(scenario.coincident_peak_baseline[norm_year][z] for z in zones)

        pos_bottom = [0.0] * len(all_years)
        neg_bottom = [0.0] * len(all_years)

        for label, data, sign, color in components:
            raw = _zone_sum_series(data, all_years, zones)
            pct = [100.0 * v / base_mw for v in raw]

            if sign == +1:
                top = [b + v for b, v in zip(pos_bottom, pct)]
                ax.fill_between(
                    x, pos_bottom, top, alpha=0.85, color=color, label=label
                )
                pos_bottom = top
            else:
                bottom = [b - v for b, v in zip(neg_bottom, pct)]
                ax.fill_between(
                    x, bottom, neg_bottom, alpha=0.85, color=color, label=label
                )
                neg_bottom = bottom

        ax.axhline(0, color="black", linewidth=0.8, linestyle="-")

        # NCP growth and coincident peak growth lines, both anchored to their
        # respective 2023-24 values and normalised by the 2023-24 CP base_mw.
        # winter_peak_baseline / coincident_peak_baseline always carry the
        # historical 2023-24 row (lower scenario files start at 2024-25).
        ncp_base_mw = sum(scenario.winter_peak_baseline["2023-24"][z] for z in zones)
        ncp_raw = _zone_sum_series(scenario.winter_peak, all_years, zones)
        ncp_growth_pct = [100.0 * (v - ncp_base_mw) / base_mw for v in ncp_raw]
        ax.plot(
            x,
            ncp_growth_pct,
            color="black",
            linewidth=1.6,
            linestyle="-",
            label="NCP growth vs 2023-24 (I-4b)",
            zorder=5,
        )

        cp_raw = _zone_sum_series(scenario.coincident_peak_baseline, all_years, zones)
        cp_growth_pct = [100.0 * (v - base_mw) / base_mw for v in cp_raw]
        ax.plot(
            x,
            cp_growth_pct,
            color="black",
            linewidth=1.6,
            linestyle="--",
            label="CP growth vs 2023-24 (I-3b)",
            zorder=5,
        )

        ex_yr = exhaustion_years.get(utility)
        if ex_yr is not None:
            ax.axvline(ex_yr, color="red", linestyle="--", linewidth=1.4, alpha=0.9)
            ax.text(
                ex_yr + 0.3,
                max(pos_bottom) * 0.97,
                f"Headroom\nexhausted\n{ex_yr}–{ex_yr - 1999:02d}",
                color="red",
                fontsize=7.5,
                va="top",
            )
        else:
            ax.text(
                0.98,
                0.97,
                "Headroom not exhausted\nwithin forecast period",
                color="red",
                fontsize=7.5,
                ha="right",
                va="top",
                transform=ax.transAxes,
            )

        ax.set_title(
            f"{utility}  (zones {', '.join(zones)},  "
            f"{norm_year} base = {base_mw:,} MW)",
            fontsize=9.5,
            loc="left",
            pad=4,
        )
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
        ax.set_ylabel(f"% of {norm_year} peak", fontsize=9)
        ax.grid(axis="y", linestyle=":", alpha=0.5)
        ax.set_xlim(x[0], x[-1])

    axes[-1].set_xlabel("Winter season starting year", fontsize=10)
    axes[-1].legend(loc="upper left", fontsize=8.5, framealpha=0.9, ncol=3)

    fig.tight_layout()
    fig.savefig(path_output, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved plot to {path_output}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--path-xlsx-baseline",
        type=Path,
        default=Path(__file__).parent
        / "2024-Gold-Book-Baseline-Forecast-Tables (2).xlsx",
        help="Path to the Gold Book baseline forecast Excel file",
    )
    parser.add_argument(
        "--path-xlsx-lower",
        type=Path,
        default=Path(__file__).parent
        / "2024-Gold-Book-Lower-Demand-Scenario-Tables.xlsx",
        help="Path to the Gold Book lower demand scenario Excel file",
    )
    parser.add_argument(
        "--scenario",
        choices=["baseline", "lower", "both"],
        default="both",
        help="Which scenario(s) to run (default: both)",
    )
    parser.add_argument(
        "--path-output-dir",
        type=Path,
        default=Path(__file__).parent,
        help="Directory for output CSV and PNG files",
    )
    args = parser.parse_args()

    scenarios_to_run: list[ScenarioData] = []
    if args.scenario in ("baseline", "both"):
        scenarios_to_run.append(load_scenario(args.path_xlsx_baseline))
    if args.scenario in ("lower", "both"):
        scenarios_to_run.append(
            load_scenario(args.path_xlsx_baseline, args.path_xlsx_lower)
        )

    for scenario in scenarios_to_run:
        s = scenario.label
        print(f"\n=== {s.upper()} SCENARIO ===")

        results = compute_headroom_exhaustion(scenario.winter_peak)

        path_csv = args.path_output_dir / f"headroom_exhaustion_year_by_utility_{s}.csv"
        with open(path_csv, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)
        print(f"Saved {len(results)} rows to {path_csv}")

        for r in results:
            print(
                f"  {r['utility']}: headroom exhausted {r['year_headroom_exhausted']}"
                f" (zones {r['zones']}, threshold {r['gb_threshold_mw']} MW)"
            )

        path_plot = args.path_output_dir / f"demand_components_by_utility_{s}.png"
        plot_demand_components_by_utility(scenario, results, path_plot)


if __name__ == "__main__":
    main()
